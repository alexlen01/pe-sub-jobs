#!/usr/bin/env python3
r"""
LP DB Export -> seed extract  (one-off, day-1 bootstrap; see pe-sub-docs/LP_DB_EXTRACT_DESIGN.md)

Standalone utility. Reads the LP DB Export (XLSX) plus a base facilities.csv and writes ALL
outputs into one directory (pe-sub-jobs/data/out/). It NEVER modifies the app tree — you review
the output, then copy the vetted CSVs into pe-sub-jobs/data/mock and run the ingest/seed jobs.

Normalization against editable reference lists in pe-sub-jobs/data/reference/ (seeded from Config):
  * Investor Type  -> supported list (investor_types.csv + investor_type_aliases.csv); unmatched
                      values are PASSED THROUGH unchanged (record kept) and dumped for review.
  * Agent LP Category (export "Classification") -> canonical Agent LP Classification
                      (agent_lp_categories.csv); unmatched passed through and dumped.
  * UBS classification: the export has no UBS class column, so it is derived from UBSAR using the
                      SAME rate-threshold tiers Run Shadow BB uses (ubs_rate_tiers.csv); the min=0
                      floor makes this total, so it always resolves (blank only if UBSAR is missing).

Outputs (always in pe-sub-jobs/data/out/):
  * lp_master.csv               - one distinct golden LP per investor_name (best-record consolidation)
  * lp_facility_seeds.csv       - per (facility, investor) LP-record seed rows
  * facilities.csv              - base + bank_status Active/Inactive + collateral_date := BBDate,
                                  PLUS a manufactured Inactive placeholder ("Unknown" bank) for any
                                  export account not in the base — so 100% of LP records seed (no rejects)
  * unmatched_investor_types.csv    - Investor Types not in the reference list (+ fuzzy suggestion)
  * unmatched_agent_categories.csv  - Agent LP Categories not in the reference list (+ suggestion)
  * EXTRACT_SUMMARY.txt         - counts + how to promote the run
  (each review report carries a leading `#` summary line and is written only when non-empty)

Design decisions this script implements (LP_DB_EXTRACT_DESIGN.md):
  D2  minimal 7-field seed; LP Size (AUM/NAV/Pension) rides in via LP Master.
  D3  facilities matched on AccountID: an existing facility whose account is in the export -> Active
      (else Inactive). An export account with NO existing facility -> a manufactured "Unknown"-bank
      Inactive placeholder, so its LP records still seed (100% insertion; no rejects).
  D4  AccountID is the join key. Real data is 1:1 (AccountID<->fund); a duplicate FndName across
      orphan accounts (a sim artifact) is disambiguated with the AccountID to keep names unique.
  D6  LP Master is cleared and repopulated with one distinct LP per investor_name, each field
      chosen by majority vote across that investor's rows (see build_master / _best).
  D7  collateral_date := BBDate  (label unchanged; wired as "Last BB Run Date").
  D8  seed ubs_default_adv_rate<-UBSAR and ubs_default_conc_limit<-UBSCL. ubs_classification is
      derived from UBSAR using the Run-Shadow-BB rate-threshold tiers (reference/ubs_rate_tiers.csv);
      always resolves (min=0 floor), blank only when UBSAR is missing.
  D10 no hard-fail: bad rows are skipped and counted; a dirty file never aborts the run.

Best-record consolidation (see build_master / _best): an investor appears on many rows (one per
facility) whose attributes may disagree. Each golden field is chosen INDEPENDENTLY by majority vote
over the non-blank values across those rows (ties -> first occurrence); blanks do not vote, so a
value present in only some rows still fills the gaps. This yields the most-agreed profile per field
rather than trusting one arbitrary row. (On the synthetic export attributes are randomized per row,
so the vote is only as meaningful as the source; a clean LP DB makes it authoritative.)

Usage (NO command-line arguments):
    1. Edit the EXPORT_FILE variable near the top of this file to point at the export to process.
    2. Run it — from any working directory:
           python pe-sub-jobs/data/lp_db_extract.py
       or, on Windows:
           powershell -ExecutionPolicy Bypass -File pe-sub-jobs/data/lp_db_extract.ps1
       or, on Linux/macOS:
           bash pe-sub-jobs/data/lp_db_extract.sh
    Re-point EXPORT_FILE and re-run for the next file. Outputs always land in pe-sub-jobs/data/out/.
"""
from __future__ import annotations

import csv
import difflib
import re
import sys
from collections import Counter, OrderedDict
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

try:
    import openpyxl
except ImportError:  # pragma: no cover
    sys.exit("openpyxl is required: pip install openpyxl")

# Everything the tool needs lives inside the pe-sub-jobs project. Paths are anchored to this file's
# location (pe-sub-jobs/data/) so the script runs from any working directory and never reaches
# outside the project.
SCRIPT_DIR  = Path(__file__).resolve().parent          # pe-sub-jobs/data/
JOBS_ROOT = SCRIPT_DIR.parent                           # pe-sub-jobs/

# ============================================================================================
#  EDIT THIS for each run — the LP DB Export to process.
#  There are NO command-line arguments: change the line below and re-run the script
#  (python pe-sub-jobs/data/lp_db_extract.py, or the .ps1/.sh wrapper). Use an absolute path, or
#  a path relative to the pe-sub-jobs project root. The default is pe-sub-jobs/data/import/ — drop
#  the export there, or repoint this line. Everything else below is a stable default.
# ============================================================================================
EXPORT_FILE = JOBS_ROOT / "data" / "import" / "LP DB Export 2026.06.25.xlsx"

# Stable defaults (rarely changed) — all within the pe-sub-jobs project:
FACILITIES_FILE = JOBS_ROOT / "data" / "mock" / "facilities.csv"  # base list, read-only
OUT_DIR = SCRIPT_DIR / "out"            # ALL outputs land here; never writes into the app tree
REFERENCE_DIR = SCRIPT_DIR / "reference"  # Config-seeded normalization lists

# --- source column order (must match the export header exactly) ----------------------------
SRC_COLS = [
    "AccountID", "FndName", "InvestorName", "Parent", "SPV", "InvestorType", "Region", "HQ",
    "InstitutionalHNW", "InvestmentGrade", "Classification", "Notes", "SP", "Moodys", "Fitch",
    "AUM", "NAV", "PensionAssets", "FundingRatio", "UBSAR", "AgentAR", "Commitments",
    "PercentOfCommitments", "Called", "Uncalled", "PercentOfUncalled", "CalledPercent",
    "AgentCL", "UBSCL", "AgentBB", "UBSBB", "BBDate",
]

# CSV header (column) orders required by the pe-sub-jobs FlatFileItemReaders.
MASTER_COLS = [
    "investor_name", "parent", "spv", "high_qty", "investor_type", "inst_vs_hnw",
    "region_location", "investment_grade", "sp", "mdy", "fitch", "aum", "nav", "pension",
    "pension_funded", "ubs_classification", "ubs_default_adv_rate", "ubs_default_conc_limit",
    "notes",
]
SEED_COLS = [
    "facility_name", "investor_name", "cap_commit", "uncalled",
    "agent_cls", "agent_rate", "agent_conc",
]
FACILITY_COLS = [
    "agent_bank", "name", "account_number", "loan_amount", "maturity_date", "bank_status",
    "bank_status_date", "ubs_participation", "collateral_date",
]


# --- formatting helpers --------------------------------------------------------------------
def _trim(dec: Decimal) -> str:
    """Fixed-point string with trailing zeros trimmed; never scientific notation."""
    s = format(dec, "f")
    if "." in s:
        s = s.rstrip("0").rstrip(".")
    return s or "0"


def blank(v) -> bool:
    return v is None or str(v).strip() == ""


def as_is(v) -> str:
    """LP-Size passthrough: keep the raw extracted value verbatim (UI does the formatting)."""
    return "" if v is None else str(v).strip()


def yn_bool(v, *, yes=("y", "yes", "true", "1")) -> str:
    return "TRUE" if str(v).strip().lower() in yes else "FALSE"


def pct(v) -> str:
    """Fraction (0.41) -> percent string ('41%'). Exact, no rounding."""
    if blank(v):
        return ""
    try:
        return _trim(Decimal(str(v)) * 100) + "%"
    except InvalidOperation:
        return ""


def dec_str(v) -> str:
    """Decimal rate/limit passthrough as a trimmed string ('0.41'), matching lp_master.csv."""
    if blank(v):
        return ""
    try:
        return _trim(Decimal(str(v)))
    except InvalidOperation:
        return str(v).strip()


def money_short(v) -> str:
    """Exact short-currency dollars ($484M, $314.6M, $8B). No rounding; falls back to a
    full grouped string only if a short form cannot represent the value losslessly."""
    if blank(v):
        return ""
    try:
        d = Decimal(str(v))
    except InvalidOperation:
        return str(v).strip()
    neg = d < 0
    d = abs(d)
    if d >= Decimal("1e9"):
        body, suffix = d / Decimal("1e9"), "B"
    elif d >= Decimal("1e6"):
        body, suffix = d / Decimal("1e6"), "M"
    elif d >= Decimal("1e3"):
        body, suffix = d / Decimal("1e3"), "K"
    else:
        body, suffix = d, ""
    return ("-" if neg else "") + "$" + _trim(body) + suffix


def iso_date(v) -> str:
    """Parse the export's M/D/YYYY BBDate to ISO (YYYY-MM-DD); '' if unparseable."""
    if blank(v):
        return ""
    s = str(v).strip()
    for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%m/%d/%y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return ""


# --- reference lists (Investor Type, Agent LP Category, UBS classification) -----------------
def _norm(s) -> str:
    """Case/punctuation-insensitive key: lowercase, non-alphanumerics collapsed to one space."""
    return re.sub(r"[^a-z0-9]+", " ", str(s or "").lower()).strip()


@dataclass
class Reference:
    itype_canonical: list[str]              # supported Investor Types (display order)
    itype_lookup: dict[str, str]            # norm(value/alias) -> canonical Investor Type
    agent_canonical: list[str]              # canonical Agent LP Categories (AGENT_CLS_OPTS)
    agent_lookup: dict[str, str]            # norm(alias) -> canonical Agent LP Category
    ubs_tiers: list[tuple[float, str]]      # (min_advance_rate_pct, classification), sorted desc


def _read_reference_rows(path: Path) -> list[list[str]]:
    """Read a reference CSV/TXT, dropping blank lines and '#'/'##' comment lines. Returns rows
    (each a list of trimmed cells). The caller decides whether the first surviving row is a header."""
    if not path.is_file():
        raise SystemExit(
            f"Reference file not found: {path}\n"
            "The pe-sub-jobs/data/reference/ lists are required. Restore it (or fix REFERENCE_DIR)."
        )
    rows: list[list[str]] = []
    with path.open(newline="", encoding="utf-8") as fh:
        for cells in csv.reader(fh):
            if not cells or not cells[0].strip() or cells[0].lstrip().startswith("#"):
                continue
            rows.append([c.strip() for c in cells])
    return rows


def load_references(ref_dir: Path) -> Reference:
    # Investor Types: canonical list + case/punct-insensitive lookup (self + aliases).
    itype_canonical = [r[0] for r in _read_reference_rows(ref_dir / "investor_types.csv")[1:]]  # drop header
    itype_lookup = {_norm(c): c for c in itype_canonical}
    alias_rows = _read_reference_rows(ref_dir / "investor_type_aliases.csv")[1:]  # drop header
    for row in alias_rows:
        if len(row) >= 2 and row[1]:
            itype_lookup[_norm(row[0])] = row[1]

    # Agent LP Categories: alias -> canonical. Canonical set = the distinct RHS values.
    agent_rows = _read_reference_rows(ref_dir / "agent_lp_categories.csv")[1:]  # drop header
    agent_lookup = {_norm(r[0]): r[1] for r in agent_rows if len(r) >= 2 and r[1]}
    agent_canonical = list(dict.fromkeys(r[1] for r in agent_rows if len(r) >= 2 and r[1]))

    # UBS classification tiers: (min_advance_rate_pct, classification), sorted highest-min first —
    # the same threshold ranges Run Shadow BB uses (ubsClassFromAgentRate). The min=0 floor makes
    # this total (every rate resolves).
    ubs_tiers: list[tuple[float, str]] = []
    for row in _read_reference_rows(ref_dir / "ubs_rate_tiers.csv")[1:]:  # drop header
        if len(row) < 2 or not row[1]:
            continue
        try:
            ubs_tiers.append((float(row[0]), row[1]))
        except ValueError:
            continue
    ubs_tiers.sort(key=lambda t: t[0], reverse=True)

    return Reference(itype_canonical, itype_lookup, agent_canonical, agent_lookup, ubs_tiers)


def _suggest(raw: str, canonical: list[str]) -> str:
    """Closest canonical value for an unmatched raw string (fuzzy hint only; never auto-applied)."""
    norms = {_norm(c): c for c in canonical}
    hit = difflib.get_close_matches(_norm(raw), list(norms), n=1, cutoff=0.6)
    return norms[hit[0]] if hit else ""


def map_investor_type(raw, ref: Reference) -> tuple[str, bool]:
    """(canonical, True) when the value maps to a supported Investor Type; otherwise the ORIGINAL
    value is passed through with (value, False) — the record is always kept, never dropped."""
    s = as_is(raw)
    if not s:
        return "", True
    canon = ref.itype_lookup.get(_norm(s))
    return (canon, True) if canon else (s, False)


def map_agent_cls(raw, ref: Reference) -> tuple[str, bool]:
    """(canonical Agent LP Category, True) when mapped; else the original value + False."""
    s = as_is(raw)
    if not s:
        return "", True
    canon = ref.agent_lookup.get(_norm(s))
    return (canon, True) if canon else (s, False)


def ubs_class_from_rate(ubsar, ref: Reference) -> tuple[str, str]:
    """Map the UBS advance rate (UBSAR) to a UBS classification using the same threshold tiers as
    Run Shadow BB: the class of the highest 'min' the rate meets or exceeds. The min=0 floor makes
    this total, so it always resolves ('mapped') unless UBSAR is missing/unparseable ('blank')."""
    if blank(ubsar):
        return "", "blank"
    try:
        rate_pct = float(ubsar) * 100
    except (TypeError, ValueError):
        return "", "blank"
    for min_pct, cls in ref.ubs_tiers:   # sorted highest-min first
        if rate_pct >= min_pct:
            return cls, "mapped"
    return "", "blank"


# --- extract -------------------------------------------------------------------------------
def read_export(path: Path, sheet: str | None = None) -> list[dict]:
    if not path.is_file():
        raise SystemExit(
            f"Export file not found: {path}\n"
            "Edit the EXPORT_FILE variable near the top of pe-sub-jobs/data/lp_db_extract.py to point at "
            "the LP DB Export .xlsx, then re-run."
        )
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet is None:
        ws = wb[wb.sheetnames[0]]
    elif sheet in wb.sheetnames:
        ws = wb[sheet]
    else:
        raise SystemExit(
            f"Sheet '{sheet}' not found in {path.name}. Available: {wb.sheetnames}"
        )
    rows_iter = ws.iter_rows(values_only=True)
    header = list(next(rows_iter))
    if header != SRC_COLS:
        raise SystemExit(
            f"Export header in '{ws.title}' does not match the expected schema.\n"
            f"  expected: {SRC_COLS}\n  found:    {header}"
        )
    return [dict(zip(SRC_COLS, r)) for r in rows_iter]


def read_facilities(path: Path) -> tuple[list[list[str]], dict[str, int]]:
    """Return (raw data rows, account_number -> row index)."""
    with path.open(newline="", encoding="utf-8") as fh:
        rows = list(csv.reader(fh))
    data = rows[1:]  # drop header (regenerated on write)
    by_acct = {}
    for i, r in enumerate(data):
        if r and r[0].strip():  # non-empty agent_bank => real facility row
            by_acct[r[2].strip()] = i
    return data, by_acct


@dataclass
class MasterResult:
    rows: list[dict]
    itype_unmatched: Counter               # raw InvestorType not in reference -> occurrences
    ubs_status: Counter                    # mapped / blank -> count


def _best(rows: list[dict], field: str) -> str:
    """Consolidate one field across all of an investor's rows into its single 'best' value:
    the MOST FREQUENTLY reported non-blank value (majority vote). Ties and all-equal-frequency
    fall back to the first occurrence, because Counter preserves insertion order. Blank/'' cells
    do not vote, so a value present in only some rows still wins over the gaps — which is exactly
    what fills sparse/typo'd real-world data. Returns '' only when every row is blank for it."""
    tally: Counter = Counter()
    for r in rows:
        v = as_is(r[field])
        if v:
            tally[v] += 1
    return tally.most_common(1)[0][0] if tally else ""


def build_master(export: list[dict], ref: Reference) -> MasterResult:
    """Build one distinct golden LP per investor_name by CONSOLIDATING all of that investor's rows
    (e.g. the ~15 occurrences across facilities). Each attribute is chosen independently by majority
    vote over the non-blank values (`_best`) — not a single arbitrary row — so the profile is the
    most-agreed value per field and gaps/typos in a minority of rows are outvoted. Investor Type is
    voted on its *canonical* mapping; UBS classification is derived from the best UBSAR via the
    Run-Shadow-BB rate tiers (rate/limit/class kept internally consistent)."""
    groups: "OrderedDict[str, list[dict]]" = OrderedDict()
    itype_seen: Counter = Counter()
    for row in export:
        name = (row["InvestorName"] or "").strip()
        if not name:
            continue
        itype_seen[as_is(row["InvestorType"])] += 1
        groups.setdefault(name, []).append(row)

    # Distinct raw Investor Types (across every row) that do not resolve to the supported list.
    itype_unmatched: Counter = Counter()
    for raw, cnt in itype_seen.items():
        if raw and not map_investor_type(raw, ref)[1]:
            itype_unmatched[raw] += cnt

    master_rows: list[dict] = []
    ubs_status: Counter = Counter()
    for name, rows in groups.items():
        # Investor Type: majority vote over the CANONICAL mapping (so "Other Institutional Investors"
        # and "Other Institutional" count together; unmatched values vote as their passed-through text).
        itype_votes: Counter = Counter()
        for r in rows:
            v = as_is(r["InvestorType"])
            if v:
                itype_votes[map_investor_type(v, ref)[0]] += 1
        investor_type = itype_votes.most_common(1)[0][0] if itype_votes else ""

        best_ubsar = _best(rows, "UBSAR")
        ubs_cls, status = ubs_class_from_rate(best_ubsar, ref)
        ubs_status[status] += 1

        master_rows.append({
            "investor_name": name,
            "parent": _best(rows, "Parent"),
            "spv": yn_bool(_best(rows, "SPV")),
            "high_qty": yn_bool(_best(rows, "HQ")),
            "investor_type": investor_type,
            "inst_vs_hnw": _best(rows, "InstitutionalHNW"),
            "region_location": _best(rows, "Region"),
            "investment_grade": yn_bool(_best(rows, "InvestmentGrade")),
            "sp": _best(rows, "SP"),
            "mdy": _best(rows, "Moodys"),
            "fitch": _best(rows, "Fitch"),
            "aum": _best(rows, "AUM"),          # LP Size passthrough (D2)
            "nav": _best(rows, "NAV"),
            "pension": _best(rows, "PensionAssets"),
            "pension_funded": pct(_best(rows, "FundingRatio")),
            "ubs_classification": ubs_cls,      # from best UBSAR via Run-Shadow-BB rate tiers
            "ubs_default_adv_rate": dec_str(best_ubsar),
            "ubs_default_conc_limit": dec_str(_best(rows, "UBSCL")),
            "notes": _best(rows, "Notes"),
        })

    return MasterResult(master_rows, itype_unmatched, ubs_status)


def build_seed(export: list[dict], name_by_acct: dict[str, str], ref: Reference,
               ) -> tuple[list[dict], Counter, Counter]:
    """Per (facility, investor) seed rows. Every export account resolves to a facility (existing or
    a manufactured placeholder), so nothing is rejected for a missing facility. Normalises the
    export's Classification (Agent LP Category) to the canonical Agent LP Classification via
    reference. Returns the seed rows, counts, and unmatched Agent LP Category values."""
    seen: set[tuple[str, str]] = set()
    seed_rows: list[dict] = []
    counts = Counter()
    agent_unmatched: Counter = Counter()      # raw Classification not in reference -> occurrences

    for row in export:
        acct = (row["AccountID"] or "").strip()
        investor = (row["InvestorName"] or "").strip()
        fac_name = name_by_acct.get(acct)
        if fac_name is None:
            # Should not happen: upsert_facilities manufactures a facility for every export account.
            counts["skipped_no_facility"] += 1
            continue
        if not investor:
            counts["skipped_blank_investor"] += 1   # investor_name is the LP key; cannot insert blank
            continue
        key = (fac_name, investor)
        if key in seen:
            counts["skipped_duplicate_pair"] += 1
            continue
        seen.add(key)

        agent_cls, matched = map_agent_cls(row["Classification"], ref)
        if not matched and agent_cls:
            agent_unmatched[agent_cls] += 1   # agent_cls holds the passed-through original here

        seed_rows.append({
            "facility_name": fac_name,
            "investor_name": investor,
            "cap_commit": money_short(row["Commitments"]),
            "uncalled": money_short(row["Uncalled"]),
            "agent_cls": agent_cls,           # normalised to Agent LP Classification (or passthrough)
            "agent_rate": pct(row["AgentAR"]),
            "agent_conc": pct(row["AgentCL"]),
        })
        counts["written"] += 1

    return seed_rows, counts, agent_unmatched


def upsert_facilities(fac_data: list[list[str]], by_acct: dict[str, int],
                      export: list[dict]) -> tuple[list[list[str]], Counter, dict[str, str]]:
    """Existing facilities: bank_status := Active if the account appears in the export else
    Inactive (Active also gets collateral_date := BBDate). Export accounts NOT in facilities.csv
    are MANUFACTURED as placeholder Inactive facilities ("Unknown" bank, name=FndName,
    account_number=AccountID, collateral_date=BBDate) so every LP record can seed — no rejects.
    Returns (rows, counts, account_number -> resolved facility name) for the seed to reuse."""
    bbdate_by_acct: "OrderedDict[str, str]" = OrderedDict()
    fnd_by_acct: "OrderedDict[str, str]" = OrderedDict()
    for row in export:
        acct = (row["AccountID"] or "").strip()
        if acct and acct not in bbdate_by_acct:
            bbdate_by_acct[acct] = iso_date(row["BBDate"])
            fnd_by_acct[acct] = as_is(row["FndName"])

    counts = Counter()
    out = [list(r) for r in fac_data]
    name_by_acct: dict[str, str] = {}
    used_norm = {_norm(r[1]) for r in out if len(r) > 1 and r[1].strip()}

    # 1) Existing facilities -> Active/Inactive.
    for acct, idx in by_acct.items():
        row = out[idx]
        name_by_acct[acct] = row[1].strip()
        if acct in bbdate_by_acct:
            row[5] = "Active"
            if bbdate_by_acct[acct]:
                row[8] = bbdate_by_acct[acct]
            counts["active"] += 1
        else:
            row[5] = "Inactive"
            counts["inactive"] += 1

    # 2) Orphan export accounts -> new placeholder Inactive facilities so their LPs still seed.
    #    A name already in use (existing facility, or another orphan sharing the FndName — a sim
    #    artifact) is disambiguated with the AccountID so each account survives as its own facility.
    for acct in bbdate_by_acct:
        if acct in by_acct:
            continue
        name = fnd_by_acct[acct] or f"Unknown Facility {acct}"
        if _norm(name) in used_norm:
            name = f"{name} ({acct})"
        used_norm.add(_norm(name))
        # agent_bank, name, account_number, loan_amount, maturity_date, bank_status,
        # bank_status_date, ubs_participation, collateral_date
        out.append(["Unknown", name, acct, "", "", "Inactive", "", "", bbdate_by_acct[acct]])
        name_by_acct[acct] = name
        counts["inactive_new"] += 1

    return out, counts, name_by_acct


def write_csv(path: Path, header: list[str], rows: list[dict]) -> None:
    with path.open("w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh, quoting=csv.QUOTE_ALL)
        w.writerow(header)
        for r in rows:
            w.writerow([r[c] for c in header])


def write_facilities(path: Path, rows: list[list[str]]) -> None:
    with path.open("w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh, quoting=csv.QUOTE_ALL)
        w.writerow(FACILITY_COLS)
        for r in rows:
            w.writerow(r[: len(FACILITY_COLS)])


def write_report(path: Path, summary: str, header: list[str], rows) -> None:
    """A review report: one `#` summary line, then a normal CSV (header + rows). The comment
    line explains what the report means so it is self-describing when opened on its own."""
    with path.open("w", newline="", encoding="utf-8") as fh:
        fh.write(f"# {summary}\n")
        w = csv.writer(fh)
        w.writerow(header)
        w.writerows(rows)


def write_summary(path: Path, lines: list[str]) -> None:
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> int:
    # No command-line arguments: the run is configured by the EXPORT_FILE variable near the top of
    # this file (edit it, then re-run). All outputs go to OUT_DIR; the app tree is never modified.
    export_path = Path(EXPORT_FILE)
    fac_path = Path(FACILITIES_FILE)
    out_dir = Path(OUT_DIR)
    ref_dir = Path(REFERENCE_DIR)
    out_dir.mkdir(parents=True, exist_ok=True)

    if not fac_path.is_file():
        raise SystemExit(f"Base facilities.csv not found: {fac_path}  (edit FACILITIES_FILE)")

    print(f"Source: {export_path}")
    ref = load_references(ref_dir)
    export = read_export(export_path)
    fac_data, by_acct = read_facilities(fac_path)

    # Facilities first: this manufactures placeholder facilities for orphan accounts and returns
    # the account -> facility name map the seed uses, so every LP record resolves to a facility.
    fac_rows, fac_counts, name_by_acct = upsert_facilities(fac_data, by_acct, export)
    mr = build_master(export, ref)
    seed_rows, seed_counts, agent_unmatched = build_seed(export, name_by_acct, ref)

    generated = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # --- app-facing CSVs (proposed; copy into pe-sub-jobs/data/mock to load) ---
    write_csv(out_dir / "lp_master.csv", MASTER_COLS, mr.rows)
    write_csv(out_dir / "lp_facility_seeds.csv", SEED_COLS, seed_rows)
    write_facilities(out_dir / "facilities.csv", fac_rows)

    # Clear any managed reports from a previous run so each run's reports reflect only this run
    # (a report is written below only when it has rows). "seed_rejects.csv" and
    # "unmatched_ubs_rate_conc.csv" are legacy — no record is rejected, and UBS classification now
    # always resolves via the rate tiers — so they are never written; remove any stale copies.
    for stale in ("lp_master_conflicts.csv", "seed_rejects.csv", "unmatched_investor_types.csv",
                  "unmatched_agent_categories.csv", "unmatched_ubs_rate_conc.csv"):
        (out_dir / stale).unlink(missing_ok=True)

    # --- review reports (only when there is something to report) ---
    if mr.itype_unmatched:
        write_report(
            out_dir / "unmatched_investor_types.csv",
            f"{len(mr.itype_unmatched)} Investor Type value(s) not in pe-sub-jobs/data/reference/"
            f"investor_types.csv. The record is still included (original value kept); an analyst "
            f"should map each to a supported type or add an alias. Generated {generated}.",
            ["raw_investor_type", "occurrences", "suggested_canonical"],
            [(v, c, _suggest(v, ref.itype_canonical)) for v, c in mr.itype_unmatched.most_common()],
        )
    if agent_unmatched:
        write_report(
            out_dir / "unmatched_agent_categories.csv",
            f"{len(agent_unmatched)} Agent LP Category value(s) (export 'Classification') not in "
            f"pe-sub-jobs/data/reference/agent_lp_categories.csv. The seed row keeps the original value; add an "
            f"alias so it maps to a canonical Agent LP Classification. Generated {generated}.",
            ["raw_classification", "occurrences", "suggested_canonical"],
            [(v, c, _suggest(v, ref.agent_canonical)) for v, c in agent_unmatched.most_common()],
        )

    # --- human summary of the whole run ---
    total_facilities = fac_counts["active"] + fac_counts["inactive"] + fac_counts["inactive_new"]
    summary_lines = [
        "LP DB Export -> seed extract — run summary",
        f"generated           : {generated}",
        f"source export       : {export_path}",
        f"base facilities.csv : {fac_path}",
        f"reference dir       : {ref_dir}",
        f"output directory    : {out_dir}",
        "",
        f"export rows read              : {len(export)}",
        f"lp_master.csv (distinct LPs)  : {len(mr.rows)}",
        f"lp_facility_seeds.csv (rows)  : {seed_counts['written']}",
        f"  skipped: duplicate pair     : {seed_counts['skipped_duplicate_pair']} (same LP in same facility)",
        f"  skipped: blank investor     : {seed_counts['skipped_blank_investor']} (no LP key - cannot insert)",
        f"  skipped: no facility        : {seed_counts['skipped_no_facility']} (should be 0)",
        f"facilities.csv                : {total_facilities} total "
        f"({fac_counts['active']} Active, {fac_counts['inactive']} Inactive, "
        f"{fac_counts['inactive_new']} new 'Unknown' placeholder Inactive)",
        "",
        "normalization:",
        f"  investor types unmatched    : {len(mr.itype_unmatched)} distinct value(s)",
        f"  agent categories unmatched  : {len(agent_unmatched)} distinct value(s)",
        f"  ubs classification (of {len(mr.rows)}) : "
        f"{mr.ubs_status['mapped']} mapped from UBSAR (rate tiers), {mr.ubs_status['blank']} blank (no UBSAR)",
        "",
        "reports (only written when non-empty):",
        f"  unmatched_investor_types.csv: {len(mr.itype_unmatched)} value(s)",
        f"  unmatched_agent_categories.csv : {len(agent_unmatched)} value(s)",
        "",
        "Every export account maps to a facility (existing or manufactured), so no LP record is",
        "rejected. To load: copy lp_master.csv, lp_facility_seeds.csv and facilities.csv from this",
        "directory into pe-sub-jobs/data/mock, then run the pe-sub-jobs ingest.",
    ]
    write_summary(out_dir / "EXTRACT_SUMMARY.txt", summary_lines)

    # --- console echo (counts onward; the header/paths block is skipped) ---
    print("\n".join(summary_lines[7:]))
    print(f"\nAll outputs written to: {out_dir}")
    print("Nothing in the app tree was modified - copy the vetted CSVs into "
          "pe-sub-jobs/data/mock to load.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
