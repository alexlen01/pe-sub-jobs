#!/usr/bin/env python3
r"""
LP DB Export -> pe-sub-jobs seed extract.

Reads two XLSX inputs from pe-sub-jobs/data/import/:
  * the LP DB Export (EXPORT_FILE) — one row per (facility account, investor)
  * AgentBankSummaryRpt.xlsx — a banded print report; the source of each facility's agent bank,
    loan amount, maturity date and agent-reported status

Writes exactly three CSVs into pe-sub-jobs/data/out/, the directory pe-sub-jobs ingests on startup:
  * lp_master.csv          one golden LP per investor_name; each field chosen independently by
                           majority vote over that investor's non-blank values across its rows
  * lp_facility_seeds.csv  one row per (facility, investor), carrying every per-LP export column
                           (facility-level AccountID/FndName/BBDate excluded)
  * facilities.csv         Agent Bank Summary rows + bank_status (Active when the account appears
                           in the export, else Inactive) + collateral_date := BBDate, plus an
                           "Unknown"-bank Inactive placeholder for every export account the report
                           does not list

Source format: the 29-column LP DB Export of 2026-08-18. Columns are located by HEADER, not by
position, so a reshuffle costs nothing here (see SRC_HEADERS). Against the previous 32-column
format that revision:
  * states the UBS LP Classification outright, where it used to be derived from the LP's attributes
  * replaces AUM / NAV / Pension Assets with one "LP Size ($ Bil)" figure plus a "LP Size Criteria"
    label naming which measure it is
  * adds Agent / UBS Excess Concentration
  * drops High Quality, Investor Type, Region, Funded Ratio and Agent Advance Rate

Normalization against the editable lists in pe-sub-jobs/data/reference/:
  * Agent LP Category (agent_lp_categories.csv) and UBS LP Classification (ubs_lp_categories.csv)
    map to canonical values; unmatched values pass through unchanged and are counted in the report
  * the agent advance rate is resolved from the row's Agent LP Category (agent_rate_map.csv),
    because the export no longer carries the rate itself, and is used as written
  * the fed UBSAR is slotted into a discrete rate group via rate_floor_map.csv (>=90 -> 90,
    75-89.9 -> 75, 65-74.9 -> 65, 50-64.9 -> 50, <50 -> 0)
  * LP Size is routed back into the aum / nav / pension_assets column its criteria names, keeping
    pe-sub-api's contract and the UI's Size Measure derivation unchanged

high_quality is no longer written at all: nothing supplies it, so pe-sub-api keeps its column on
the schema default rather than being fed a fabricated value. investor_type, region_location and
funding_ratio stay in the CSV header but go out blank, which pe-sub-api reads as "not resubmitted"
and therefore leaves any existing LP Master value intact.

The input is dirty by design (name drift, 'A minus' ratings, unit mix-ups, LP Size range strings),
so the parsers are tolerant and no row aborts the run. Every export account resolves to a facility,
existing or manufactured, so lp_facility_seeds.csv retains 100% of the export's records; the run
prints those counts and nothing else.

Usage (no command-line arguments): edit EXPORT_FILE below, then run from any working directory:
    python pe-sub-jobs/scripts/lp_db_extract.py
"""
from __future__ import annotations

import csv
import re
import sys
from collections import Counter, OrderedDict
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

try:
    import openpyxl
except ImportError:  # pragma: no cover
    sys.exit("openpyxl is required: pip install openpyxl")

# Paths are anchored to this file's location so the script runs from any working directory.
SCRIPT_DIR = Path(__file__).resolve().parent            # pe-sub-jobs/scripts/
JOBS_ROOT = SCRIPT_DIR.parent                           # pe-sub-jobs/
DATA_DIR = JOBS_ROOT / "data"

# ============================================================================================
#  EDIT THIS for each run — the LP DB Export to process. Absolute, or relative to pe-sub-jobs/.
# ============================================================================================
EXPORT_FILE = DATA_DIR / "import" / "LP DB Export 2026.08.18.xlsx"

AGENT_BANK_SUMMARY_FILE = DATA_DIR / "import" / "AgentBankSummaryRpt.xlsx"
OUT_DIR = DATA_DIR / "out"              # all outputs land here
REFERENCE_DIR = DATA_DIR / "reference"  # normalization lists

# --- source columns -------------------------------------------------------------------------
# Internal names for the export's 29 columns, in the order the 2026-08-18 format lists them.
# Columns are located BY HEADER, not by position (see SRC_HEADERS / read_export), so a further
# reshuffle needs no code change - only a new spelling needs one.
SRC_COLS = [
    "AccountID", "FndName", "InvestorName", "Parent", "SPV", "UbsClassification",
    "InstitutionalHNW", "InvestmentGrade", "Classification", "SP", "Moodys", "Fitch",
    "LpSizeBil", "LpSizeCriteria", "Commitments", "Uncalled", "UBSAR", "AgentCL", "UBSCL",
    "PercentOfCommitments", "Called", "PercentOfUncalled", "CalledPercent",
    "AgentExcessConc", "UBSExcessConc", "AgentBB", "UBSBB", "Notes", "BBDate",
]

# Accepted header spellings per column. Matching runs through _norm(), which lowercases and
# collapses every run of non-alphanumerics to one space - so it absorbs the format's own quirks
# without needing an entry for each: the embedded CRLF in "LP Size\r\n($ Bil)", the "Insitutional"
# typo, the "Moody'S" capitalisation, and "Investment Grade?"'s trailing question mark.
#
# The lists deliberately cover THREE vocabularies at once, because all three land on this reader:
#   * the current LP DB Export headers (first entry of each list);
#   * the pre-2026-08-18 export's terse headers (FndName, InvestorName, UBSAR, ...), so an archived
#     workbook still parses for the columns it does have;
#   * the readable headers the platform's own LP Records export writes
#     (pe-sub-ui/src/services/lpExportService.ts), so a workbook exported from the UI can be fed
#     straight back in. Keep those in step with that file.
SRC_HEADERS = {
    "AccountID":            ["AccountID", "Account ID"],
    "FndName":              ["FndName", "Fund Name", "Facility Name"],
    "InvestorName":         ["Investor Name", "InvestorName"],
    "Parent":               ["Parent", "Parent / Sponsor"],
    "SPV":                  ["SPV", "SPV Flag"],
    "UbsClassification":    ["UBS LP Classification", "UBS (Internal) LP Classification",
                             "UBS Classification", "UBS LP Category"],
    "InstitutionalHNW":     ["Insitutional vs HNW", "Institutional vs HNW", "InstitutionalHNW"],
    "InvestmentGrade":      ["Investment Grade?", "Investment Grade", "InvestmentGrade"],
    "Classification":       ["Agent LP Classification", "Classification", "Agent LP Category"],
    "SP":                   ["S&P", "SP", "S&P Rating"],
    "Moodys":               ["Moody'S", "Moodys", "Moody's Rating"],
    "Fitch":                ["Fitch", "Fitch Rating"],
    "LpSizeBil":            ["LP Size ($ Bil)", "LP Size", "LpSizeBil"],
    "LpSizeCriteria":       ["LP Size Criteria", "Size Measure", "Size Metric Type"],
    "Commitments":          ["Capital Commitments", "Commitments"],
    "Uncalled":             ["Uncalled Capital", "Uncalled"],
    "UBSAR":                ["UBS Advance Rate", "UBSAR", "UBS Advance Rate (%)"],
    "AgentCL":              ["Agent Concentration Limit", "AgentCL"],
    "UBSCL":                ["UBS Concentration Limit", "UBSCL"],
    "PercentOfCommitments": ["% of Capital Commitments", "% of Commitments", "PercentOfCommitments"],
    "Called":               ["Called Capital", "Called"],
    "PercentOfUncalled":    ["% of Uncalled Capital", "PercentOfUncalled"],
    "CalledPercent":        ["% of LP Called", "CalledPercent"],
    "AgentExcessConc":      ["Agent Excess Concentration", "Agent Excess Conc Base"],
    "UBSExcessConc":        ["UBS Excess Concentration", "UBS Excess Conc Base"],
    "AgentBB":              ["Agent Borrowing Base", "AgentBB"],
    "UBSBB":                ["UBS Borrowing Base", "UBSBB"],
    "Notes":                ["Notes"],
    "BBDate":               ["BBDate", "Collateral Date", "BB Date"],
}

# Columns the 2026-08-18 format dropped, kept here only so a stale workbook is diagnosed with a
# useful message instead of a bare "missing column" list. None of them feed the outputs any more:
#   InvestorType / Region / HQ  - HQ is gone from the platform feed entirely; investor type and
#                                 region stay in the schema but are governed outside this feed.
#   AUM / NAV / PensionAssets   - superseded by LP Size ($ Bil) + LP Size Criteria.
#   FundingRatio                - no longer sourced.
#   AgentAR                     - resolved from the Agent LP Classification (agent_rate_map.csv).
RETIRED_HEADERS = {
    "InvestorType": "Investor Type", "Region": "Region / Location", "HQ": "High Quality",
    "AUM": "AUM", "NAV": "NAV", "PensionAssets": "Pension Assets",
    "FundingRatio": "Funded Ratio (%)", "AgentAR": "Agent Advance Rate (%)",
}

# --- numeric normalization ------------------------------------------------------------------
# Two vocabularies reach this reader with the SAME headers but differently shaped values. The LP DB
# Export writes rates and shares as fractions (0.154); the platform's own LP Records export writes
# them for a spreadsheet, as percent strings or bare percent numbers ("15.4%", 94) with money as
# display strings ("$428,800,000"). Four headers are identical in both - "% of Uncalled Capital",
# "% of LP Called", "Agent Concentration Limit", "UBS Concentration Limit" - so the header cannot
# tell them apart and a header-keyed rule would divide the export's own fractions by 100.
#
# The shape of the VALUE decides instead, which is unambiguous and idempotent: a share or rate is a
# fraction by definition, so anything carrying a '%' or exceeding 1 is a percent and is scaled down,
# while 0.154 is already correct and left alone.
PERCENT_COLS = {"UBSAR", "PercentOfCommitments", "PercentOfUncalled", "CalledPercent"}
MONEY_COLS = {"Commitments", "Called", "Uncalled", "AgentBB", "UBSBB",
              "AgentExcessConc", "UBSExcessConc"}
# A concentration limit is either a percent of uncalled ("7.5%") or an absolute cap ("$25,000,000")
# in the same column - the '%' sign is what tells them apart; bare numbers pass through for the
# downstream magnitude split to resolve.
LIMIT_COLS = {"AgentCL", "UBSCL"}


def _parsed_number(v) -> "tuple[Decimal, bool] | None":
    """Strip the display formatting off one cell -> (number, carried_a_percent_sign).
    '$428,800,000' -> (428800000, False) - '7.5%' -> (7.5, True) - 94 -> (94, False).
    None when the cell holds nothing numeric, in which case the caller keeps it verbatim."""
    s = str(v).strip().replace(",", "").replace("$", "")
    was_pct = s.endswith("%")
    if was_pct:
        s = s[:-1].strip()
    try:
        return Decimal(s), was_pct
    except InvalidOperation:
        return None


def normalize_numeric(col: str, v):
    """Bring one numeric cell to the feed's internal shape: fractions for rates and shares, plain
    dollars for money. Unparseable values pass through untouched, same as any other dirty cell."""
    if blank(v) or col not in (PERCENT_COLS | MONEY_COLS | LIMIT_COLS):
        return v
    parsed = _parsed_number(v)
    if parsed is None:
        return v                                   # unparseable: pass through
    number, was_pct = parsed
    if col in PERCENT_COLS:                        # '15.4%' or 94 -> 0.154 / 0.94; 0.154 stays
        return _trim(number / 100 if was_pct or abs(number) > 1 else number)
    if col in LIMIT_COLS:                          # '7.5%' -> 0.075; '$25,000,000' -> 25000000
        return _trim(number / 100 if was_pct else number)
    return _trim(number)                           # money: '$428,800,000' -> 428800000

# CSV header (column) orders required by the pe-sub-jobs FlatFileItemReaders.
# high_quality is gone: the export no longer carries it, and nothing else supplies it. The platform
# keeps its own column on the schema default (TRUE) rather than being fed a fabricated value.
# investor_type, region_location and funding_ratio stay in the header but go out BLANK from this
# feed - the 2026-08-18 format dropped them and they are governed elsewhere (LP Master is a
# bank-wide store; investor type and ratings are analyst-compiled from Pitchbook and the agencies).
# Keeping the columns means pe-sub-api's contract is unchanged and a value already on an LP Master
# record is not clobbered by this feed.
MASTER_COLS = [
    "investor_name", "parent", "spv", "investor_type", "institutional_or_hnw",
    "region_location", "investment_grade", "sp_rating", "moodys_rating", "fitch_rating", "aum", "nav", "pension_assets",
    "funding_ratio", "ubs_lp_category", "ubs_default_advance_rate", "ubs_default_concentration_limit",
    "notes",
]
SEED_COLS = [
    "facility_name", "investor_name", "capital_commitment", "uncalled_capital",
    "agent_lp_category", "agent_advance_rate", "agent_concentration_limit",
    "parent", "spv", "investor_type", "institutional_or_hnw", "region_location",
    "investment_grade", "ubs_lp_category", "sp_rating", "moodys_rating", "fitch_rating", "aum", "nav", "pension_assets",
    "funding_ratio", "pct_of_fund_commitments", "called_capital", "pct_of_fund_uncalled", "pct_lp_called",
    "ubs_concentration_limit", "ubs_advance_rate", "agent_excess_concentration",
    "ubs_excess_concentration", "agent_borrowing_base", "ubs_borrowing_base", "notes",
]
FACILITY_COLS = [
    "agent_bank", "name", "account_number", "loan_amount", "maturity_date", "bank_status",
    "bank_status_date", "ubs_participation", "collateral_date",
]

# Agent Bank Summary column layout (must match the report header exactly). Index 4 is an unnamed
# spacer holding the report's subtotal amounts.
ABS_COLS = [
    "Agent", "Borrower", "AccountNumber", "LoanAmount", "", "MaturityDate",
    "FacilityStatus", "FacilityStatusDate",
]
ABS_TOTAL_MARKER = "accesstotalsloanamount"  # _norm() prefix of the subtotal / grand-total rows


# --- formatting helpers --------------------------------------------------------------------
def _trim(dec: Decimal) -> str:
    """Fixed-point string with trailing zeros trimmed; never scientific notation."""
    s = format(dec, "f")
    if "." in s:
        s = s.rstrip("0").rstrip(".")
    return s or "0"


def blank(v) -> bool:
    return v is None or str(v).strip() == ""


def as_is(v) -> str:
    """Passthrough: the raw extracted value verbatim, trimmed."""
    return "" if v is None else str(v).strip()


def yn_bool(v, *, yes=("y", "yes", "true", "1")) -> str:
    return "TRUE" if str(v).strip().lower() in yes else "FALSE"


def pct(v) -> str:
    """Fraction (0.41) -> percent string ('41%'). Exact, no rounding."""
    if blank(v):
        return ""
    try:
        return _trim(Decimal(str(v)) * 100) + "%"
    except InvalidOperation:
        return ""


def dec_str(v) -> str:
    """Decimal rate/limit passthrough as a trimmed string ('0.41')."""
    if blank(v):
        return ""
    try:
        return _trim(Decimal(str(v)))
    except InvalidOperation:
        return str(v).strip()


def money_short(v) -> str:
    """Exact short-currency dollars ($484M, $314.6M, $8B); no rounding."""
    if blank(v):
        return ""
    try:
        d = Decimal(str(v))
    except InvalidOperation:
        return str(v).strip()
    neg = d < 0
    d = abs(d)
    if d >= Decimal("1e9"):
        body, suffix = d / Decimal("1e9"), "B"
    elif d >= Decimal("1e6"):
        body, suffix = d / Decimal("1e6"), "M"
    elif d >= Decimal("1e3"):
        body, suffix = d / Decimal("1e3"), "K"
    else:
        body, suffix = d, ""
    return ("-" if neg else "") + "$" + _trim(body) + suffix


def iso_date(v) -> str:
    """Parse a feed date to ISO (YYYY-MM-DD); '' if unparseable. The export carries text dates;
    the Agent Bank Summary carries real Excel dates, which openpyxl returns as datetime."""
    if blank(v):
        return ""
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    s = str(v).strip()
    for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%m/%d/%y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return ""


# --- reference lists ------------------------------------------------------------------------
def _norm(s) -> str:
    """Case/punctuation-insensitive key: lowercase, non-alphanumerics collapsed to one space."""
    return re.sub(r"[^a-z0-9]+", " ", str(s or "").lower()).strip()


@dataclass
class Reference:
    agent_lookup: dict[str, str]            # norm(alias) -> canonical Agent LP Category
    ubs_lookup: dict[str, str]              # norm(alias) -> canonical UBS LP Classification
    agent_rates: dict[str, float]           # canonical Agent LP Category -> advance rate percent
    rate_floors: list[tuple[float, float]]  # (min_rate_pct, group_pct), sorted highest-min first


def _read_reference_rows(path: Path) -> list[list[str]]:
    """Read a reference CSV, dropping blank lines and '#' comment lines."""
    if not path.is_file():
        raise SystemExit(
            f"Reference file not found: {path}\n"
            "The pe-sub-jobs/data/reference/ lists are required. Restore it (or fix REFERENCE_DIR)."
        )
    rows: list[list[str]] = []
    with path.open(newline="", encoding="utf-8") as fh:
        for cells in csv.reader(fh):
            if not cells or not cells[0].strip() or cells[0].lstrip().startswith("#"):
                continue
            rows.append([c.strip() for c in cells])
    return rows


def load_references(ref_dir: Path) -> Reference:
    # investor_types.csv / investor_type_aliases.csv are no longer read: the 2026-08-18 export
    # dropped the Investor Type column, so there is nothing here to normalize. The files stay in
    # data/reference/ because they mirror classification_config.INVESTOR_TYPE_OPTS for the platform.
    agent_rows = _read_reference_rows(ref_dir / "agent_lp_categories.csv")[1:]
    agent_lookup = {_norm(r[0]): r[1] for r in agent_rows if len(r) >= 2 and r[1]}

    ubs_rows = _read_reference_rows(ref_dir / "ubs_lp_categories.csv")[1:]
    ubs_lookup = {_norm(r[0]): r[1] for r in ubs_rows if len(r) >= 2 and r[1]}

    # Agent advance rate by Agent LP Category: the export stopped carrying an Agent Advance Rate
    # column, so the rate is resolved from the row's category instead of read from the feed.
    agent_rates: dict[str, float] = {}
    for row in _read_reference_rows(ref_dir / "agent_rate_map.csv")[1:]:
        if len(row) < 2:
            continue
        try:
            agent_rates[row[0]] = float(str(row[1]).rstrip("%"))
        except ValueError:
            continue

    # Floor Map: a rate takes the group of the highest 'min' it meets or exceeds; the min=0 floor
    # makes it total.
    rate_floors: list[tuple[float, float]] = []
    for row in _read_reference_rows(ref_dir / "rate_floor_map.csv")[1:]:
        if len(row) < 2:
            continue
        try:
            rate_floors.append((float(row[0]), float(row[1])))
        except ValueError:
            continue
    rate_floors.sort(key=lambda t: t[0], reverse=True)

    return Reference(agent_lookup, ubs_lookup, agent_rates, rate_floors)


def map_agent_cls(raw, ref: Reference) -> tuple[str, bool]:
    """(canonical Agent LP Category, True) when mapped; else the original value with False."""
    s = as_is(raw)
    if not s:
        return "", True
    canon = ref.agent_lookup.get(_norm(s))
    return (canon, True) if canon else (s, False)


# --- UBS classification ---------------------------------------------------------------------
# The 2026-08-18 export states the UBS LP Classification outright, so it is no longer derived here.
# The waterfall this replaced inferred it from agency ratings, pension assets, NAV, AUM and the
# HNW/SPV flags - four of which the format no longer carries - so deriving it is neither possible
# nor wanted: the LP DB is the system of record for this field.
def map_ubs_cls(raw, ref: Reference) -> tuple[str, bool]:
    """(canonical UBS LP Classification, True) when the fed value maps to one of the nine classes;
    otherwise the ORIGINAL value with (value, False). The record is always kept - an unrecognised
    class is reported, never dropped or silently rewritten."""
    s = as_is(raw)
    if not s:
        return "", True
    canon = ref.ubs_lookup.get(_norm(s))
    return (canon, True) if canon else (s, False)


# --- LP size ---------------------------------------------------------------------------------
# "LP Size ($ Bil)" + "LP Size Criteria" replace the old AUM / NAV / PensionAssets trio. The
# criteria column names which measure the figure is, using the same vocabulary as the platform's
# LP_SIZE_CRITERIA_OPTS ("AUM", "NAV", "Assets"), so the value is routed back into whichever of the
# three schema columns it belongs to. That keeps pe-sub-api's contract and the UI's Size Measure
# derivation (aum ? 'AUM' : nav ? 'NAV' : pension_assets ? 'Assets') working unchanged.
# Keyed by _norm(), so case and punctuation are already absorbed. The spelled-out variants are here
# because the criteria cell is analyst-typed free text in practice ("Total AUM", "Net Asset Value"),
# and an unrecognised label costs the row its whole LP Size - the figure has no meaning without a
# basis to attribute it to.
SIZE_CRITERIA_COL = {
    "aum": "aum", "total aum": "aum", "assets under management": "aum",
    "nav": "nav", "net asset value": "nav", "fund nav": "nav",
    "assets": "pension_assets", "total assets": "pension_assets",
    "pension assets": "pension_assets", "pension": "pension_assets",
}

_SIZE_UNIT_MULT = {"": 1.0, "k": 1e-6, "m": 1e-3, "mn": 1e-3, "mm": 1e-3,
                   "b": 1.0, "bn": 1.0, "t": 1e3, "tn": 1e3, "trn": 1e3}


def parse_size_bil(v) -> float | None:
    """Tolerant parser for the LP Size column -> billions of dollars. The column's unit is $Bn, so a
    bare number is already billions ('13.5' -> 13.5) - unlike the old AUM/NAV free text, where a
    bare number meant absolute dollars. An explicit unit still wins when the analyst typed one
    ('$13.5B', '900M', '1.2 bn', '>5B'), and a range takes its LOW end ('5-8' -> 5).
    None when nothing numeric reads."""
    if blank(v):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().lower().replace(",", "").replace("$", "")
    s = s.lstrip("<>~ ").rstrip("+ ")
    s = re.sub(r"\.(?!\d)", " ", s)          # stray dots ('1.33. bn') - keep decimal points only
    parts = re.findall(r"(\d+(?:\.\d+)?)\s*([a-z]*)", s)
    if not parts:
        return None
    units = [u for _, u in parts if u in _SIZE_UNIT_MULT and u]
    default_unit = units[-1] if units else ""   # '5-8bn': the shared unit applies to both ends
    vals = []
    for num, unit in parts:
        mult = _SIZE_UNIT_MULT.get(unit) if unit in _SIZE_UNIT_MULT else _SIZE_UNIT_MULT.get(default_unit, 1.0)
        vals.append(float(num) * mult)
    return min(vals) if vals else None


def size_display(v) -> str:
    """One LP Size cell as the short-currency display string the aum/nav/pension_assets columns hold
    ('13.5' -> '$13.5B'). Unparseable text passes through verbatim, so a dirty cell stays visible
    for review instead of becoming a wrong number."""
    bil = parse_size_bil(v)
    if bil is None:
        return as_is(v)
    return money_short(Decimal(str(bil)) * Decimal("1e9"))


def size_columns(row: dict) -> dict[str, str]:
    """Route one row's LP Size into the aum / nav / pension_assets column its criteria names.
    An unrecognised or blank criteria leaves all three blank rather than guessing a measure -
    the figure without its basis is not attributable to any of them."""
    out = {"aum": "", "nav": "", "pension_assets": ""}
    col = SIZE_CRITERIA_COL.get(_norm(row["LpSizeCriteria"]))
    if col:
        out[col] = size_display(row["LpSizeBil"])
    return out


# --- agent advance rate ----------------------------------------------------------------------
def agent_rate_frac(raw_category, ref: Reference) -> float | None:
    """The agent advance rate for a row, as a fraction. The export no longer carries the rate, so it
    is resolved from the row's canonical Agent LP Category via agent_rate_map.csv. None when the
    category is blank or unrecognised - a made-up rate would be indistinguishable from a fed one."""
    canon, matched = map_agent_cls(raw_category, ref)
    if not matched or not canon:
        return None
    pct_value = ref.agent_rates.get(canon)
    return None if pct_value is None else pct_value / 100


def rate_group_pct(v, ref: Reference) -> float | None:
    """Floor Map: slot a raw advance-rate fraction (0.93) into its rate group in percent (90.0) —
    the group of the highest 'min' the rate meets or exceeds. None when missing/unparseable."""
    if blank(v):
        return None
    try:
        rate_pct = float(v) * 100
    except (TypeError, ValueError):
        return None
    for min_pct, group_pct in ref.rate_floors:   # sorted highest-min first
        if rate_pct >= min_pct:
            return group_pct
    return None


def floor_rate_pct(v, ref: Reference) -> str:
    """Floor-mapped rate as a percent string ('90%'); '' when unmapped."""
    group = rate_group_pct(v, ref)
    return "" if group is None else _trim(Decimal(str(group))) + "%"


def floor_rate_frac(v, ref: Reference) -> str:
    """Floor-mapped rate as a fraction string ('0.9'); '' when unmapped."""
    group = rate_group_pct(v, ref)
    return "" if group is None else _trim(Decimal(str(group)) / 100)


# --- extract -------------------------------------------------------------------------------
def read_export(path: Path, sheet: str | None = None) -> list[dict]:
    if not path.is_file():
        raise SystemExit(
            f"Export file not found: {path}\n"
            "Edit the EXPORT_FILE variable near the top of pe-sub-jobs/scripts/lp_db_extract.py "
            "to point at the LP DB Export .xlsx, then re-run."
        )
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet is None:
        ws = wb[wb.sheetnames[0]]
    elif sheet in wb.sheetnames:
        ws = wb[sheet]
    else:
        raise SystemExit(
            f"Sheet '{sheet}' not found in {path.name}. Available: {wb.sheetnames}"
        )
    rows_iter = ws.iter_rows(values_only=True)
    header = ["" if h is None else str(h) for h in next(rows_iter)]
    # Columns are addressed by NAME, not position, so the 2026-08-18 reshuffle needed no change
    # here and a further one will not either. Matching goes through _norm(), which lowercases and
    # collapses runs of non-alphanumerics - that is what absorbs the header quirks the format ships
    # with, notably the embedded CRLF in "LP Size\r\n($ Bil)", the "Insitutional" typo and the
    # trailing "?" on "Investment Grade?". Anything unrecognised is ignored, so an extra trailing
    # column never breaks the read.
    header_to_col = {_norm(alias): col for col, aliases in SRC_HEADERS.items() for alias in aliases}
    column_at: dict[str, int] = {}
    for i, name in enumerate(header):
        col = header_to_col.get(_norm(name))
        if col is not None:
            column_at.setdefault(col, i)
    missing = [c for c in SRC_COLS if c not in column_at]
    if missing:
        # A stale pre-2026-08-18 workbook fails on exactly the five added columns, so name them and
        # say which retired ones are present - far more useful than a bare missing-column list.
        found_retired = [h for h in header
                         if _norm(h) in {_norm(v) for v in RETIRED_HEADERS.values()}]
        hint = ""
        if found_retired:
            hint = (f"\n  This looks like a PRE-2026-08-18 export: it still carries "
                    f"{found_retired}, which the current format dropped. Re-export it, or run an "
                    f"older revision of this script against it.")
        raise SystemExit(
            f"""Export header in '{ws.title}' is missing {len(missing)} of the {len(SRC_COLS)} expected columns: {missing}
  Accepted header spellings per column are listed in SRC_HEADERS near the top of this script.
  found: {header}{hint}"""
        )
    rows = []
    for r in rows_iter:
        row = {c: (r[column_at[c]] if column_at[c] < len(r) else None) for c in SRC_COLS}
        # Rates/shares arrive as fractions from the LP DB Export and as percents from the platform's
        # own export under several identical headers; normalize_numeric decides on the value, not
        # the header, and is idempotent for values already in the feed's shape.
        for c in PERCENT_COLS | MONEY_COLS | LIMIT_COLS:
            row[c] = normalize_numeric(c, row[c])
        rows.append(row)
    return rows


def read_agent_bank_summary(path: Path) -> tuple[list[list[str]], dict[str, int]]:
    """Read the Agent Bank Summary report into FACILITY_COLS-shaped rows.
    Returns (rows, account_number -> row index).

    The report is a banded print layout, not a flat table:
      * the agent bank appears once on its own group-header row (Agent set, Borrower blank) and is
        carried down onto the facility rows beneath it, which leave the Agent cell blank;
      * each group ends with an 'AccessTotalsLoanAmount:' subtotal row carrying no facility.

    Dirty rows are absorbed rather than fatal: a repeated (AccountNumber, Borrower) pair is a
    reprint and is dropped; a Borrower name already taken by a different account is suffixed with
    its AccountID; one AccountNumber listed against two borrowers keeps the FIRST row as the owner
    of the LP join.

    ubs_participation and collateral_date are not in the report — collateral_date is filled from
    the export's BBDate by upsert_facilities. bank_status comes from FacilityStatus here and is
    then overridden by upsert_facilities with the export-match result."""
    if not path.is_file():
        raise SystemExit(
            f"Agent Bank Summary report not found: {path}\n"
            "It is the source of every facility's agent bank, loan amount, maturity date and "
            "agent-reported status. Drop it in pe-sub-jobs/data/import/, or edit the "
            "AGENT_BANK_SUMMARY_FILE variable near the top of this script, then re-run."
        )
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows_iter = ws.iter_rows(values_only=True)
    header = [as_is(c) for c in next(rows_iter)]
    if header[: len(ABS_COLS)] != ABS_COLS:
        raise SystemExit(
            f"Agent Bank Summary header in '{ws.title}' does not match the expected schema.\n"
            f"  expected: {ABS_COLS}\n  found:    {header}"
        )

    data: list[list[str]] = []
    by_acct: dict[str, int] = {}
    seen_pair: set[tuple[str, str]] = set()
    used_norm: set[str] = set()
    agent = ""

    for raw in rows_iter:
        cells = (list(raw) + [None] * len(ABS_COLS))[: len(ABS_COLS)]
        text = [as_is(c) for c in cells]
        if not any(text):
            continue
        if _norm(text[3]).startswith(ABS_TOTAL_MARKER):   # subtotal / grand-total band
            continue
        if text[0] and not text[1]:                       # agent group header -> carry down
            agent = text[0]
            continue
        name, acct = text[1], text[2]
        if not name or not acct:
            continue
        if (acct, _norm(name)) in seen_pair:              # reprint of a row already taken
            continue
        seen_pair.add((acct, _norm(name)))
        if _norm(name) in used_norm:
            name = f"{name} ({acct})"
        used_norm.add(_norm(name))
        # The row's own Agent cell wins if the report fills it; otherwise the carried-down header.
        # "Unknown" satisfies FacilityRowProcessor's non-blank agent_bank rule.
        data.append([text[0] or agent or "Unknown", name, acct, text[3], iso_date(cells[5]),
                     text[6], iso_date(cells[7]), "", ""])
        by_acct.setdefault(acct, len(data) - 1)

    return data, by_acct


def facility_key(row: dict) -> str:
    """Identity of the facility an export row belongs to. AccountID is that identity - a fund name
    can be reported under two accounts, and those are two different facilities.

    A row with no AccountID cannot be joined on account, so it falls back to its fund name under a
    prefix no real account number can produce. Such rows group into one placeholder facility per
    fund name instead of being discarded: the export is the system of record, and a missing account
    number is a data-quality problem to report, not grounds for dropping an LP."""
    acct = (row["AccountID"] or "").strip()
    return acct or f"?fnd:{_norm(as_is(row['FndName']))}"


def _recent_first(rows: list[dict]) -> list[dict]:
    """An investor's rows ordered by BB run date, most recent first. A later submission supersedes
    an earlier one, so this ordering - not a headcount - decides every consolidated attribute.

    sorted() is stable and stays stable under reverse=True, so rows sharing a BBDate keep their
    export order and the earliest-listed of them wins. Rows with no BBDate sort last: an undated
    submission never outranks a dated one."""
    return sorted(rows, key=lambda r: iso_date(r["BBDate"]) or "", reverse=True)


def _latest(rows: list[dict], field: str) -> str:
    """Consolidate one field across an investor's rows, which MUST already be _recent_first():
    the value from the most recent submission that actually reported it.

    A blank on a newer row means 'not resubmitted this cycle', not 'cleared' - ratings and
    concentration limits are routinely omitted from a submission that still carries a current
    investor type - so the search falls through to the next-most-recent non-blank value rather
    than letting a blank erase a known one. '' only when every row is blank."""
    for r in rows:
        v = as_is(r[field])
        if v:
            return v
    return ""


def build_master(export: list[dict], ref: Reference) -> list[dict]:
    """One golden LP per investor_name, consolidated by recency: each attribute is taken from the
    most recent BB run that reported it (see _recent_first / _latest). A later submission both
    improves on data quality and carries the attributes that are current - investor type, LP
    category, ratings - so it supersedes earlier ones field by field rather than being outvoted by
    a mass of older rows.

    Investor Type is mapped to canonical AFTER the recency pick, so an older row cannot supply the
    label. UBS classification is derived from the recency-consolidated attributes."""
    groups: "OrderedDict[str, list[dict]]" = OrderedDict()
    for row in export:
        name = (row["InvestorName"] or "").strip()
        if not name:
            continue
        groups.setdefault(name, []).append(row)

    master_rows: list[dict] = []
    for name, rows in groups.items():
        rows = _recent_first(rows)
        # LP Size and its criteria are consolidated as a PAIR from the same submission: a figure
        # taken from one BB run and a measure label from another would mislabel the number. The
        # search falls through to the next-most-recent submission that actually RESOLVES - a row
        # whose criteria drifted to a label no alias covers yields nothing, and stopping there would
        # blank the LP's size even though an older row states it perfectly well.
        size = {"aum": "", "nav": "", "pension_assets": ""}
        for r in rows:
            routed = size_columns(r)
            if any(routed.values()):
                size = routed
                break

        master_rows.append({
            "investor_name": name,
            "parent": _latest(rows, "Parent"),
            "spv": yn_bool(_latest(rows, "SPV")),
            # investor_type / region_location / funding_ratio left blank: the 2026-08-18 export
            # dropped them and no other column implies them. Blank means "not resubmitted", which
            # pe-sub-api treats as "keep what LP Master already holds" - not as a clearing edit.
            "investor_type": "",
            "institutional_or_hnw": _latest(rows, "InstitutionalHNW"),
            "region_location": "",
            "investment_grade": yn_bool(_latest(rows, "InvestmentGrade")),
            "sp_rating": _latest(rows, "SP"),
            "moodys_rating": _latest(rows, "Moodys"),
            "fitch_rating": _latest(rows, "Fitch"),
            "aum": size["aum"],
            "nav": size["nav"],
            "pension_assets": size["pension_assets"],
            "funding_ratio": "",
            "ubs_lp_category": map_ubs_cls(_latest(rows, "UbsClassification"), ref)[0],
            "ubs_default_advance_rate": floor_rate_frac(_latest(rows, "UBSAR"), ref),
            "ubs_default_concentration_limit": dec_str(_latest(rows, "UBSCL")),
            "notes": _latest(rows, "Notes"),
        })

    return master_rows


@dataclass
class SeedResult:
    rows: list[dict]
    counts: Counter
    anomalies: list[str]


def build_seed(export: list[dict], name_by_acct: dict[str, str], ref: Reference) -> SeedResult:
    """One seed row per export row, carrying every per-LP export column. NOTHING is dropped:
    the export is the system of record and every LP it lists must reach the platform.

    AccountID - not facility name - is the identity of a facility here. The same fund name can be
    reported under two accounts, and those are two different facilities, so a repeated
    (facility name, investor) pair is legitimate and is written out. upsert_facilities guarantees
    each account owns a DISTINCT facility name (colliding names are suffixed with their AccountID),
    which is what lets these rows stay resolvable downstream, where LpRecordsSeedJobConfig looks a
    facility up by facilityName alone.

    Two conditions are still recorded as anomalies - a blank investor name, and the same
    (AccountID, investor) appearing twice - but their rows are written anyway and reported by main()
    so they surface at ingest instead of disappearing here.

    Classification and Investor Type are normalised (unmatched passed through); ubs_lp_category is
    derived per row from that row's own attributes."""
    seen: set[tuple[str, str]] = set()
    seed_rows: list[dict] = []
    counts = Counter()
    anomalies: list[str] = []

    for row in export:
        acct = (row["AccountID"] or "").strip()
        investor = (row["InvestorName"] or "").strip()
        # Every key is in the map: upsert_facilities manufactures a placeholder facility for every
        # facility_key the Agent Bank Summary does not report, so this cannot fail.
        fkey = facility_key(row)
        fac_name = name_by_acct[fkey]
        if not investor:
            counts["blank_investor"] += 1
            anomalies.append(f"blank investor name on account {acct or '(no account)'}")
        if not acct:
            counts["blank_account"] += 1
            anomalies.append(f"no AccountID for investor {investor!r} "
                             f"(placed in {fac_name!r} by fund name)")
        key = (fkey, investor)
        if key in seen:
            counts["duplicate_account_investor"] += 1
            anomalies.append(f"account {acct or '(no account)'} lists investor "
                             f"{investor!r} more than once")
        seen.add(key)

        agent_cls, agent_matched = map_agent_cls(row["Classification"], ref)
        ubs_cls, ubs_matched = map_ubs_cls(row["UbsClassification"], ref)
        if not agent_matched:
            counts["unmatched_agent_category"] += 1
        if not ubs_matched:
            counts["unmatched_ubs_classification"] += 1
        if not as_is(row["UbsClassification"]):
            counts["blank_ubs_classification"] += 1
        size = size_columns(row)
        agent_rate = agent_rate_frac(row["Classification"], ref)
        if agent_rate is None:
            counts["unresolved_agent_rate"] += 1

        seed_rows.append({
            "facility_name": fac_name,
            "investor_name": investor,
            "capital_commitment": money_short(row["Commitments"]),
            "uncalled_capital": money_short(row["Uncalled"]),
            "agent_lp_category": agent_cls,
            # Resolved from the Agent LP Category, NOT floor-mapped. The floor map exists to tame a
            # dirty fed rate (0.93 -> the 90 group); this value comes from agent_rate_map.csv, which
            # mirrors config, so it is already canonical. Flooring it would silently turn Designated
            # Institutional's configured 60% into 50%, since 60 is not one of the 90/75/65/50/0
            # groups the UBS side uses.
            "agent_advance_rate": pct(agent_rate),
            "agent_concentration_limit": pct(row["AgentCL"]),
            "parent": as_is(row["Parent"]),
            "spv": yn_bool(row["SPV"]),
            "investor_type": "",
            "institutional_or_hnw": as_is(row["InstitutionalHNW"]),
            "region_location": "",
            "investment_grade": yn_bool(row["InvestmentGrade"]),
            "ubs_lp_category": ubs_cls,
            "sp_rating": as_is(row["SP"]),
            "moodys_rating": as_is(row["Moodys"]),
            "fitch_rating": as_is(row["Fitch"]),
            "aum": size["aum"],
            "nav": size["nav"],
            "pension_assets": size["pension_assets"],
            "funding_ratio": "",
            "pct_of_fund_commitments": pct(row["PercentOfCommitments"]),
            "called_capital": money_short(row["Called"]),
            "pct_of_fund_uncalled": pct(row["PercentOfUncalled"]),
            "pct_lp_called": pct(row["CalledPercent"]),
            "ubs_concentration_limit": pct(row["UBSCL"]),
            "ubs_advance_rate": floor_rate_pct(row["UBSAR"], ref),
            "agent_excess_concentration": money_short(row["AgentExcessConc"]),
            "ubs_excess_concentration": money_short(row["UBSExcessConc"]),
            "agent_borrowing_base": money_short(row["AgentBB"]),
            "ubs_borrowing_base": money_short(row["UBSBB"]),
            "notes": as_is(row["Notes"]),
        })
        counts["written"] += 1

    return SeedResult(seed_rows, counts, anomalies)


def upsert_facilities(fac_data: list[list[str]], by_acct: dict[str, int],
                      export: list[dict]) -> tuple[list[list[str]], dict[str, str]]:
    """Facilities from the Agent Bank Summary: bank_status := Active if the account appears in the
    export else Inactive (Active also gets collateral_date := the account's most recent BBDate),
    overriding the report's own FacilityStatus.

    An export account the report does not list is NOT a reason to drop its LPs. It is manufactured
    as a placeholder Inactive facility carrying the three things the export does know - name=FndName
    (Facility Name), account_number=AccountID, collateral_date=last BB date when non-blank - with
    "Unknown" as the agent bank, so every LP record still seeds and the facility is visibly
    unconfirmed rather than absent.

    Returns (rows, facility_key -> resolved facility name)."""
    # An account's Last BB date is the MOST RECENT BB run across its rows, not the first one listed -
    # the export is not guaranteed to arrive in date order, and the same recency rule that governs
    # lp_master applies here. Insertion order stays the export's first-seen order, so the
    # placeholder facilities below keep a stable, reproducible ordering.
    bbdate_by_acct: "OrderedDict[str, str]" = OrderedDict()
    fnd_by_acct: "OrderedDict[str, str]" = OrderedDict()
    acctno_by_key: "OrderedDict[str, str]" = OrderedDict()
    for row in export:
        key = facility_key(row)
        bbdate = iso_date(row["BBDate"])
        if key not in bbdate_by_acct:
            bbdate_by_acct[key] = bbdate
            fnd_by_acct[key] = as_is(row["FndName"])
            acctno_by_key[key] = (row["AccountID"] or "").strip()
        else:
            if bbdate > bbdate_by_acct[key]:    # ISO dates compare lexicographically; '' loses
                bbdate_by_acct[key] = bbdate
            if not fnd_by_acct[key]:
                fnd_by_acct[key] = as_is(row["FndName"])

    out = [list(r) for r in fac_data]
    name_by_acct: dict[str, str] = {}
    used_norm = {_norm(r[1]) for r in out if len(r) > 1 and r[1].strip()}

    # Reported facilities. Status is resolved per row off that row's own account; only the owning
    # row (by_acct) lends its name to the LP seeds.
    for acct, idx in by_acct.items():
        name_by_acct[acct] = out[idx][1].strip()
    for row in out:
        acct = row[2].strip()
        if acct in bbdate_by_acct:
            row[5] = "Active"
            if bbdate_by_acct[acct]:
                row[8] = bbdate_by_acct[acct]
        else:
            row[5] = "Inactive"

    # Orphan export accounts -> placeholder Inactive facilities, name disambiguated by AccountID
    # when already in use, so each account survives as its own facility. A key with no AccountID
    # (see facility_key) lands here too, with a blank account_number - FacilityRowProcessor accepts
    # that, and it is the only way such a row reaches the platform at all.
    for key in bbdate_by_acct:
        if key in by_acct:
            continue
        acctno = acctno_by_key[key]
        name = fnd_by_acct[key] or (f"Unknown Facility {acctno}" if acctno
                                    else "Unknown Facility (no account)")
        if _norm(name) in used_norm:
            name = f"{name} ({acctno})" if acctno else f"{name} (no account)"
        used_norm.add(_norm(name))
        out.append(["Unknown", name, acctno, "", "", "Inactive", "", "", bbdate_by_acct[key]])
        name_by_acct[key] = name

    return out, name_by_acct


def write_csv(path: Path, header: list[str], rows: list[dict]) -> None:
    with path.open("w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh, quoting=csv.QUOTE_ALL)
        w.writerow(header)
        for r in rows:
            w.writerow([r[c] for c in header])


def write_facilities(path: Path, rows: list[list[str]]) -> None:
    with path.open("w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh, quoting=csv.QUOTE_ALL)
        w.writerow(FACILITY_COLS)
        for r in rows:
            w.writerow(r[: len(FACILITY_COLS)])


def main() -> int:
    export_path = Path(EXPORT_FILE)
    abs_path = Path(AGENT_BANK_SUMMARY_FILE)
    out_dir = Path(OUT_DIR)
    ref_dir = Path(REFERENCE_DIR)
    out_dir.mkdir(parents=True, exist_ok=True)

    ref = load_references(ref_dir)
    # Small file first, so a missing/renamed report fails before the big export is parsed.
    fac_data, by_acct = read_agent_bank_summary(abs_path)
    export = read_export(export_path)

    # Facilities first: manufactures placeholders for orphan accounts and returns the
    # account -> facility name map the seed uses, so every LP record resolves to a facility.
    fac_rows, name_by_acct = upsert_facilities(fac_data, by_acct, export)
    master_rows = build_master(export, ref)
    sr = build_seed(export, name_by_acct, ref)

    # Clear data/out/ so it holds only this run's three CSVs. Done after all inputs are read, so a
    # failed read leaves the last good outputs in place. import/ and reference/ are never touched.
    for old in out_dir.iterdir():
        if old.is_file():
            old.unlink()

    write_csv(out_dir / "lp_master.csv", MASTER_COLS, master_rows)
    write_csv(out_dir / "lp_facility_seeds.csv", SEED_COLS, sr.rows)
    write_facilities(out_dir / "facilities.csv", fac_rows)

    # Retention is an invariant, not a metric: every export row must appear in
    # lp_facility_seeds.csv. A mismatch is a bug in this script, so it fails the run.
    reported = sum(1 for r in fac_rows if r[5] == "Active")
    print(f"export rows            : {len(export)}")
    print(f"lp_facility_seeds rows : {sr.counts['written']}")
    print(f"lp_master rows         : {len(master_rows)} (one per distinct investor name)")
    print(f"facilities             : {len(fac_rows)} "
          f"({reported} active from the report, {len(fac_rows) - reported} inactive)")

    # Normalization outcomes. These are not failures - the value is written through unchanged - but
    # a non-zero count means a reference list is behind the feed and should be topped up.
    norm_counts = [
        ("unmatched Agent LP Category", sr.counts["unmatched_agent_category"],
         "agent_lp_categories.csv"),
        ("unmatched UBS classification", sr.counts["unmatched_ubs_classification"],
         "ubs_lp_categories.csv"),
        ("blank UBS classification", sr.counts["blank_ubs_classification"],
         "fed empty by the export"),
        ("unresolved agent advance rate", sr.counts["unresolved_agent_rate"],
         "agent_rate_map.csv"),
    ]
    if any(n for _, n, _ in norm_counts):
        print()
        print("normalization           : values written through unchanged, listed to be fixed at source")
        for label, n, where in norm_counts:
            if n:
                print(f"  {label:<30}: {n}  ({where})")

    if sr.anomalies:
        # Written, not dropped - listed so they can be dealt with at source.
        print()
        print(f"anomalies              : {len(sr.anomalies)} row(s) written but flagged")
        print(f"  blank investor name          : {sr.counts['blank_investor']}")
        print(f"  missing AccountID            : {sr.counts['blank_account']}")
        print(f"  repeated (account, investor) : {sr.counts['duplicate_account_investor']}")
        for line in sr.anomalies[:20]:
            print(f"    - {line}")
        if len(sr.anomalies) > 20:
            print(f"    ... and {len(sr.anomalies) - 20} more")

    if sr.counts["written"] != len(export):
        raise SystemExit(
            f"RETENTION FAILURE: {len(export)} export rows produced "
            f"{sr.counts['written']} seed rows. Every export row must be written; "
            "the outputs above are incomplete and must not be ingested."
        )
    print()
    print(f"retained               : 100% ({len(export)}/{len(export)} rows)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
