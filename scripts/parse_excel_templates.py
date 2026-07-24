#!/usr/bin/env python3
r"""
Parse Excel workbook → generate BB template

Analyzes an Excel workbook and generates a template metadata file (BB-Template-Import-<slug>.xlsx).
Standalone utility with no dependencies on external services or repositories. Features robust
investor-data validation to skip Legend, Notes, and metadata rows.

USAGE
    python parse_excel_templates.py <workbook.xlsx>

Example:
    python parse_excel_templates.py agent-bb-2026.xlsx

Output:
    Writes BB-Template-Import-<slug>.xlsx to the same directory as the input file.
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from dataclasses import dataclass, field, asdict
from datetime import datetime
from pathlib import Path
from typing import Optional

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:  # pragma: no cover
    sys.exit("openpyxl is required: pip install openpyxl")

# ── Analysis heuristic constants ──
MIN_HEADER_MATCHES_DEFAULT = 3     # Minimum columns matching field dictionary to recognize a header row
HEADER_SCAN_ROWS = 150             # Maximum rows to scan for header row detection
MAX_GROUPS = 12                    # Maximum LP-category groups to detect per sheet
MAX_ROWS_PER_SHEET_DEFAULT = 10_000  # Maximum rows per sheet to analyze (safety cap)
MIN_DATA_ROWS_AFTER_HEADER = 2     # Minimum data rows with 60% density required to confirm a sheet is a grid
DATA_ROW_SCAN_LIMIT = 30           # Rows to scan after header for data validation

AGENT_LABELS = {
    "agent bank", "agent", "administrative agent", "administered by", "prepared by", "lender",
}
# Banner row keywords (normalized for matching)
SKIP_BANNER_NORMS = ("total", "subtotal", "sub total", "grand total", "sum", "net total")
# Default skip keywords to mark rows that should be excluded from data processing
DEFAULT_SKIP_KEYWORDS = ["Total", "Subtotal", "Sub-Total", "Grand Total", "Sum", "Net Total"]

# Metadata/non-data row detection: keywords indicating a row is explanatory, legend, or footer.
METADATA_ROW_KEYWORDS = {
    "legend", "notes", "footer", "source", "disclaimer", "footnote", "key", "explanation",
    "shading", "color", "formatting", "indicates", "denotes", "represents", "symbol",
    "blue text", "green ", "yellow ", "underlined", "bold", "italic", "font", "style",
    "means", "indicate", "shows", "mean", "marked", "marked as",
}
# Minimum percentage of matched columns that must contain data for a row to be considered "investor data"
MIN_INVESTOR_DATA_THRESHOLD = 0.60  # 60% of matched columns must have data

TAB_ROLES = ("LP_GRID", "CONCENTRATION", "CAPITAL_CALL", "TOP_SHEET")  # bb_template_tabs.tab_role CHECK
TEMPLATE_CLASSES = ("A", "B", "C")                                     # BbTemplateRequest @Pattern

# Built-in field mappings: canonical field names and their common aliases.
# These are the core fields used for template analysis.
BUNDLED_FALLBACK_FIELDS: dict[str, list[str]] = {
    "Investor Name": ["Investor Name", "Investor", "LP Name", "Limited Partner"],
    "Fund Sleeve": ["Fund Sleeve", "Fund Vehicle", "Feeder Fund", "Feeder"],
    "Parent / Sponsor": ["Parent / Sponsor", "Parent", "Sponsor"],
    "SPV Flag": ["SPV Flag", "SPV", "Special Purpose Vehicle"],
    "Region / Location": ["Region / Location", "Region", "Location", "Domicile"],
    "Investor Type": ["Investor Type", "Investor Classification", "Category of Investor"],
    "Institutional vs HNW": ["Institutional vs HNW", "HNW Flag", "Investor Segment"],
    "LP Category": ["LP Category", "Agent LP Classification", "LP Classification"],
    "Investment Grade Flag": ["Investment Grade Flag", "Investment Grade", "IG Flag"],
    "S&P Rating": ["S&P", "S&P Rating"],
    "Moody's Rating": ["Moody's", "Moody's Rating", "Moodys"],
    "Fitch Rating": ["Fitch", "Fitch Rating"],
    "Capital Commitments": ["Capital Commitments", "Committed Capital", "Total Commitment"],
    "% of Capital Commitments": ["% of Capital Commitments", "% Commitment"],
    "Called Capital": ["Called Capital", "Drawn Capital", "Funded Capital"],
    "Uncalled Capital": ["Uncalled Capital", "Unfunded Commitment", "Remaining Commitment"],
    "% of Uncalled Capital": ["% of Uncalled Capital", "% Uncalled", "Uncalled %"],
    "AUM": ["AUM", "Assets Under Management", "Total AUM"],
    "NAV": ["NAV", "Net Asset Value"],
    "Pension Assets": ["Pension Assets", "Pension Fund Assets"],
    "Borrowing Base": ["Borrowing Base", "Agent Borrowing Base", "Agent BB"],
    "Advance Rate": ["Advance Rate", "Agent Advance Rate", "Adv. Rate", "Rate"],
    "Concentration Limit": ["Concentration Limit", "Agent Concentration Limit", "Conc. Limit"],
    "Excess Concentration": ["Excess Concentration", "Conc. Overage"],
    "Notes": ["Notes", "Comments", "Remarks"],
}


# ==============================================================================================
#  Small self-describing value types (mirror TemplateProposal's shapes; see docstring above)
# ==============================================================================================
@dataclass
class ProfiledField:
    """Field value with confidence level and supporting evidence."""
    value: object
    confidence: str   # "high" | "medium" | "low"
    evidence: str

    @staticmethod
    def of(value, confidence: str, evidence: str) -> "ProfiledField":
        return ProfiledField(value, confidence, evidence)


@dataclass
class ProposedGroup:
    header_text: str
    classification: str
    confidence: str
    row_index: int              # 0-based row this banner was found on (diagnostic only)


@dataclass
class SkipRowHit:
    row_index: int              # 0-based
    matched_keyword: str
    sample_text: str


@dataclass
class ColorFlag:
    style: str                  # e.g. "Fill #FFFF00" / "Font #0000FF"
    sample_cell: str            # e.g. "C42"
    occurrences: int


@dataclass
class TabAnalysis:
    sheet_name: str
    tab_sort: int
    header: ProfiledField                       # value = 0-based row index, or None if no grid found
    header_row_span: int
    columns: list[str] = field(default_factory=list)
    matched_canonical: int = 0
    matched_fields: list[dict] = field(default_factory=list)     # [{header, canonical, lpMasterField}]
    unmatched_headers: list[str] = field(default_factory=list)   # never silently dropped — listed
    groups: list[ProposedGroup] = field(default_factory=list)
    skip_row_keywords: list[str] = field(default_factory=list)
    skip_rows_detected: list[SkipRowHit] = field(default_factory=list)
    color_flags: list[ColorFlag] = field(default_factory=list)
    structure_fingerprint: tuple = field(default_factory=tuple)
    duplicate_of: Optional[str] = None
    notes: list[str] = field(default_factory=list)


@dataclass
class WorkbookAnalysis:
    file_name: str
    slug: ProfiledField = None
    agent_name: ProfiledField = None
    title: ProfiledField = None
    template_class: ProfiledField = None
    workbook_shape: ProfiledField = None         # "single" | "multiple" | "none"
    tabs: list[TabAnalysis] = field(default_factory=list)
    non_grid_sheets: list[str] = field(default_factory=list)
    auto_discover_tabs: ProfiledField = None
    detect_keys: list[str] = field(default_factory=list)
    overall_confidence: str = "low"
    dictionary_source: str = ""
    dictionary_field_count: int = 0
    truncated_sheets: list[str] = field(default_factory=list)
    analyzed_at: str = ""


# ==============================================================================================
#  Field Mapping Dictionary — built-in only
# ==============================================================================================
@dataclass
class Dictionary:
    canonical_fields: list[str]           # ordered, display order
    alias_lookup: dict[str, str]          # normalize(alias) -> canonical field
    lp_master_field: dict[str, str]       # canonical field -> lp_master_field (may be "")
    source: str                           # "bundled"


_WS_RE = re.compile(r"\s+")


def _normalize(s) -> str:
    """Lowercase, non-alnum -> space, whitespace collapsed."""
    if s is None:
        return ""
    return _WS_RE.sub(" ", re.sub(r"[^a-z0-9\s]", " ", str(s).lower())).strip()


def load_dictionary() -> Dictionary:
    """Load built-in field mappings from BUNDLED_FALLBACK_FIELDS."""
    alias_lookup = {_normalize(a): canon for canon, aliases in BUNDLED_FALLBACK_FIELDS.items() for a in aliases}
    for canon in BUNDLED_FALLBACK_FIELDS:
        alias_lookup[_normalize(canon)] = canon
    return Dictionary(list(BUNDLED_FALLBACK_FIELDS), alias_lookup, {}, source="bundled")


# ==============================================================================================
#  ExcelAnalyzer — structural + semantic analysis of a raw Agent BB workbook
# ==============================================================================================
class ExcelAnalyzer:
    """Detects LP-grid sheets, header rows, column->canonical mappings, LP-category group banners,
    and skip-row candidates in an Excel workbook. Never raises on a single malformed sheet/row —
    flags and continues."""

    def __init__(self, dictionary: Dictionary, *, min_header_matches: int = MIN_HEADER_MATCHES_DEFAULT,
                 max_rows_per_sheet: int = MAX_ROWS_PER_SHEET_DEFAULT, color_scan: bool = True):
        self.dictionary = dictionary
        self.min_header_matches = min_header_matches
        self.max_rows_per_sheet = max_rows_per_sheet
        self.color_scan = color_scan

    # ── Workbook-level entry point ──────────────────────────────────────────────────────────
    def analyze_workbook(self, path: Path, *, agent_bank_override: Optional[str] = None,
                          slug_override: Optional[str] = None,
                          template_class_override: Optional[str] = None) -> WorkbookAnalysis:
        read_only = not self.color_scan   # color/merged-cell inspection needs style access
        wb = openpyxl.load_workbook(path, read_only=read_only, data_only=True)

        analysis = WorkbookAnalysis(file_name=path.name, analyzed_at=datetime.now().isoformat(timespec="seconds"))
        analysis.dictionary_source = self.dictionary.source
        analysis.dictionary_field_count = len(self.dictionary.canonical_fields)

        first_grid_sheet_rows: Optional[list[list]] = None
        first_grid_header_row = -1
        seen_fingerprints: dict[tuple, str] = {}
        tab_sort = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows, truncated = self._materialize_rows(ws)
            if truncated:
                analysis.truncated_sheets.append(sheet_name)

            scan = self._find_header_row(rows)
            if scan is None:
                analysis.non_grid_sheets.append(sheet_name)
                continue

            header_row_idx, matched_count = scan
            header_cells = self._non_blank(rows[header_row_idx]) if header_row_idx < len(rows) else []
            matched_fields, unmatched = self._map_columns(header_cells)

            # Validate that sheet has actual investor data rows, not just a header match
            matched_column_headers = [f["header"] for f in matched_fields]
            if not self._has_valid_data_rows(rows, header_row_idx, matched_column_headers):
                analysis.non_grid_sheets.append(sheet_name)
                continue

            # Validate data rows post-header: detect sparse/metadata rows
            metadata_row_indices, additional_skip_hits = self._validate_data_rows(
                rows, header_row_idx, matched_column_headers
            )

            tab_sort += 1
            fingerprint = tuple(sorted(_normalize(c) for c in header_cells))
            duplicate_of = seen_fingerprints.get(fingerprint)
            if duplicate_of is None:
                seen_fingerprints[fingerprint] = sheet_name

            # Detect groups, but exclude metadata rows from consideration
            groups = self._detect_groups(rows, header_row_idx, excluded_rows=set(metadata_row_indices))
            skip_rows = self._detect_skip_rows(rows, header_row_idx)
            # Merge additional skip rows detected by data validation
            skip_rows.extend(additional_skip_hits)
            # Deduplicate by row index
            skip_rows_by_idx = {hit.row_index: hit for hit in skip_rows}
            skip_rows = list(skip_rows_by_idx.values())
            skip_rows.sort(key=lambda x: x.row_index)

            tab = TabAnalysis(
                sheet_name=sheet_name,
                tab_sort=tab_sort,
                header=ProfiledField.of(header_row_idx, "high",
                                         f"alias-scored header row ({matched_count} canonical matches, threshold {self.min_header_matches})"),
                header_row_span=self._detect_header_span(ws, header_row_idx, len(header_cells), read_only),
                columns=header_cells,
                matched_canonical=matched_count,
                matched_fields=matched_fields,
                unmatched_headers=unmatched,
                groups=groups,
                skip_row_keywords=list(DEFAULT_SKIP_KEYWORDS),
                skip_rows_detected=skip_rows,
                structure_fingerprint=fingerprint,
                duplicate_of=duplicate_of,
            )
            if self.color_scan and not read_only:
                tab.color_flags = self._detect_color_flags(ws, header_row_idx, len(rows))
            if duplicate_of:
                tab.notes.append(
                    f"Column structure identical to sheet '{duplicate_of}' — confirm whether this is a "
                    f"genuine continuation/sleeve (keep both, in tabSort order) or an accidental duplicate export."
                )
            if unmatched:
                tab.notes.append(
                    f"{len(unmatched)} header(s) did not match any canonical field: {unmatched}. "
                    f"Confirm these columns are intentionally out of scope for this template."
                )
            analysis.tabs.append(tab)

            if first_grid_sheet_rows is None:
                first_grid_sheet_rows = rows
                first_grid_header_row = header_row_idx

        n_tabs = len(analysis.tabs)
        analysis.workbook_shape = ProfiledField.of(
            "multiple" if n_tabs > 1 else "single",   # BbTemplateRequest has no third state; a 0-tab
                                                        # proposal is caught separately (see main()) —
                                                        # this value is a placeholder in that case, but
                                                        # its LOW confidence below says so honestly.
            "medium" if n_tabs > 1 else ("high" if n_tabs == 1 else "low"),
            f"{n_tabs} grid sheet(s) detected out of {len(wb.sheetnames)} total sheet(s)",
        )
        analysis.agent_name = self._infer_agent(first_grid_sheet_rows, first_grid_header_row, agent_bank_override)
        analysis.title = self._infer_title(first_grid_sheet_rows)
        slug = (slug_override or "").strip() or self._slug_from_filename(path.name)
        analysis.slug = ProfiledField.of(
            slug, "high" if slug_override else ("high" if slug else "low"),
            "operator-provided" if slug_override else "derived from filename",
        )
        analysis.template_class = ProfiledField.of(
            (template_class_override or "A"),
            "high" if template_class_override else "low",
            "operator-provided" if template_class_override
            else "default 'A' — template_class is a business/workflow classification, not a "
                 "structural signal; confirm with the template owner",
        )
        analysis.auto_discover_tabs = ProfiledField.of(
            False, "low",
            "assumed named tabs — confirm if this agent's tab names vary run-to-run (would need auto-discover)",
        )
        analysis.detect_keys = self._detect_keys(slug, analysis.title.value)
        analysis.overall_confidence = (
            "low" if not analysis.tabs
            else "high" if analysis.tabs[0].matched_canonical >= self.min_header_matches
            else "medium"
        )
        return analysis

    # ── Sheet materialization ───────────────────────────────────────────────────────────────
    def _materialize_rows(self, ws) -> tuple[list[list], bool]:
        """Reads all cell values into a plain list of lists (openpyxl read_only worksheets stream
        from XML — a single Python list at BB-workbook scale, hundreds of rows, is trivial after
        that; this trades a little memory for being able to re-scan the sheet freely below without
        juggling generator re-entrancy). Capped at max_rows_per_sheet as a large-file safety net."""
        rows: list[list] = []
        truncated = False
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= self.max_rows_per_sheet:
                truncated = True
                break
            rows.append(["" if c is None else c for c in row])
        return rows, truncated

    def _non_blank(self, cells: list) -> list[str]:
        return [str(c).strip() for c in cells if c is not None and str(c).strip() != ""]

    # ── Header detection (TemplateProfiler.findHeaderRow / matchesAlias) ───────────────────
    def _find_header_row(self, rows: list[list]) -> Optional[tuple[int, int]]:
        best_row, best_matched = -1, 0
        limit = min(HEADER_SCAN_ROWS, len(rows))
        for i in range(limit):
            matched = sum(1 for cell in rows[i] if cell and self._matches_alias(_normalize(cell)))
            if matched > best_matched:
                best_matched, best_row = matched, i
        return (best_row, best_matched) if best_matched >= self.min_header_matches else None

    def _has_valid_data_rows(self, rows: list[list], header_row_idx: int,
                            matched_columns: list[str], min_data_rows: int = MIN_DATA_ROWS_AFTER_HEADER) -> bool:
        """Check if sheet has at least min_data_rows with 60%+ data density in matched columns.
        This ensures a sheet is a real investor grid, not just a summary/reference sheet
        with incidental header matches."""
        if not matched_columns or header_row_idx < 0:
            return False

        matched_col_count = len(matched_columns)
        min_data_cols = max(1, int(matched_col_count * MIN_INVESTOR_DATA_THRESHOLD))

        valid_data_rows = 0
        limit = min(len(rows), header_row_idx + 1 + DATA_ROW_SCAN_LIMIT)

        for i in range(header_row_idx + 1, limit):
            row_cells = rows[i]
            if not row_cells:
                continue

            # Skip rows flagged as metadata
            populated = self._non_blank(row_cells)
            if not populated:
                continue

            row_text = " ".join(populated).lower()
            is_metadata = any(kw in row_text for kw in METADATA_ROW_KEYWORDS)
            if is_metadata:
                continue

            # Count data in matched columns only
            data_col_count = sum(1 for j in range(min(matched_col_count, len(row_cells)))
                               if row_cells[j] and str(row_cells[j]).strip())

            if data_col_count >= min_data_cols:
                valid_data_rows += 1
                if valid_data_rows >= min_data_rows:
                    return True

        return False

    def _matches_alias(self, norm_cell: str) -> bool:
        if not norm_cell:
            return False
        if norm_cell in self.dictionary.alias_lookup:
            return True
        for alias_norm in self.dictionary.alias_lookup:
            if not alias_norm:
                continue
            if len(alias_norm) >= 4 and " " in alias_norm and alias_norm in norm_cell:
                return True
            if len(norm_cell) >= 4 and " " in norm_cell and norm_cell in alias_norm:
                return True
        return False

    def _map_columns(self, header_cells: list[str]) -> tuple[list[dict], list[str]]:
        matched, unmatched = [], []
        for h in header_cells:
            norm = _normalize(h)
            canonical = self.dictionary.alias_lookup.get(norm)
            if canonical is None:
                for alias_norm, canon in self.dictionary.alias_lookup.items():
                    if alias_norm and ((len(alias_norm) >= 4 and " " in alias_norm and alias_norm in norm)
                                        or (len(norm) >= 4 and " " in norm and norm in alias_norm)):
                        canonical = canon
                        break
            if canonical:
                matched.append({"header": h, "canonical": canonical,
                                 "lpMasterField": self.dictionary.lp_master_field.get(canonical, "")})
            else:
                unmatched.append(h)
        return matched, unmatched

    # ── Group / LP-category banner detection (TemplateProfiler.detectGroups/classifyGroup) ──
    def _detect_groups(self, rows: list[list], header_row: int, excluded_rows: Optional[set] = None) -> list[ProposedGroup]:
        """Detect LP-category group banners (single-cell rows indicating investor classification).
        Exclude rows that have been flagged as metadata/sparse by _validate_data_rows."""
        groups: list[ProposedGroup] = []
        excluded_rows = excluded_rows or set()
        for i in range(header_row + 1, len(rows)):
            if len(groups) >= MAX_GROUPS:
                break
            if i in excluded_rows:
                # Skip rows that were flagged as metadata/sparse
                continue
            populated = self._non_blank(rows[i])
            if len(populated) != 1:
                continue
            text = populated[0]
            norm = _normalize(text)
            if not (3 <= len(norm) <= 60) or self._is_numeric(text) or any(b in norm for b in SKIP_BANNER_NORMS):
                continue
            classification, confidence = self._classify_group(norm, text)
            groups.append(ProposedGroup(text, classification, confidence, i))
        return groups

    def _classify_group(self, norm: str, verbatim: str) -> tuple[str, str]:
        if "exclud" in norm or "ineligible" in norm:
            return "Ineligible Investors", "medium"
        if "designated" in norm and "pwm" in norm:
            return "Designated PWM", "medium"
        if "designated" in norm or "institutional" in norm:
            return "Designated Institutional", "medium"
        if ("non" in norm and "rated" in norm) or "unrated" in norm:
            return "Non-Rated Included", "medium"
        if "rated" in norm or "includ" in norm:
            return "Rated Included", "medium"
        return verbatim, "low"   # verbatim self-map (e.g. a sleeve/feeder section header)

    # ── Row-skip detection — broader than banner-only: any row after the header whose FIRST
    #    populated cell contains a skip keyword, so a subtotal line embedded mid-grid (label +
    #    numeric columns still populated) is caught, not just a lone-cell banner. ──────────────
    def _detect_skip_rows(self, rows: list[list], header_row: int) -> list[SkipRowHit]:
        hits: list[SkipRowHit] = []
        for i in range(header_row + 1, len(rows)):
            populated = self._non_blank(rows[i])
            if not populated:
                continue
            first_norm = _normalize(populated[0])
            hit_kw = next((kw for kw in DEFAULT_SKIP_KEYWORDS if _normalize(kw) in first_norm), None)
            if hit_kw:
                hits.append(SkipRowHit(i, hit_kw, " | ".join(populated[:4])))
        return hits

    # ── Validate data rows post-header: detect sparse/metadata rows that are not investor data ──
    def _validate_data_rows(self, rows: list[list], header_row_idx: int,
                           matched_columns: list[str]) -> tuple[list[int], list[SkipRowHit]]:
        """
        Post-header row validation: automatically detect and flag rows that are sparse, metadata-only,
        or junk, rather than actual investor data. Returns (metadata_row_indices, additional_skip_hits).

        Heuristics:
        1. Detect metadata keywords (Legend, Notes, Footer, Shading, etc.) in any cell — HIGH PRIORITY
        2. For rows with multiple populated cells, check % of matched columns with data
        3. Detect rows with all same value repeated (formatting junk)
        4. Detect rows with ONLY single explanatory words

        Single-cell rows are generally legitimate group headers (Rated Included, Non-Rated, etc.),
        so we only flag them if they contain explicit metadata keywords. This prevents false positives
        on legitimate group classifications while catching Legend/Notes rows.

        This is invoked after header detection to enrich skip_rows and prevent Legend/Notes from
        being classified as groups.
        """
        metadata_rows: list[int] = []
        additional_skips: list[SkipRowHit] = []

        if not matched_columns or header_row_idx < 0:
            return metadata_rows, additional_skips

        matched_col_count = len(matched_columns)
        if matched_col_count == 0:
            return metadata_rows, additional_skips

        # Minimum columns that must have data for this to be "investor data" row
        min_data_cols = max(1, int(matched_col_count * MIN_INVESTOR_DATA_THRESHOLD))

        limit = min(len(rows), header_row_idx + 1 + 2000)  # scan up to 2000 rows after header

        for i in range(header_row_idx + 1, limit):
            row_cells = rows[i]
            if not row_cells:
                continue

            # Extract populated cells (non-blank strings)
            populated = self._non_blank(row_cells)
            if not populated:
                continue

            row_text = " ".join(populated).lower()
            is_metadata_row = False
            matched_keyword = None

            # Heuristic 1 (HIGH PRIORITY): Check if any cell contains metadata keywords
            # This catches Legend, Notes, Footer, Shading, etc.
            for kw in METADATA_ROW_KEYWORDS:
                if kw in row_text:
                    is_metadata_row = True
                    matched_keyword = kw
                    break

            # Heuristic 2 (MODERATE PRIORITY): Check data density across matched columns
            # Single-cell rows are typically group headers or banners, so don't apply sparse check
            # Only apply to rows with multiple populated cells (which should have more investor data)
            is_sparse = False
            if not is_metadata_row and len(populated) > 1:
                data_col_count = 0
                for j in range(min(matched_col_count, len(row_cells))):
                    cell = row_cells[j]
                    if cell and str(cell).strip():
                        data_col_count += 1
                # Flag as sparse if low data density across matched columns
                if data_col_count < min_data_cols:
                    is_sparse = True

            # Heuristic 3: Check if row is all the same value repeated (formatting artifact)
            is_repetitive = False
            if not is_metadata_row and len(populated) >= 3:
                first_val = _normalize(populated[0])
                if first_val and all(_normalize(v) == first_val for v in populated):
                    is_repetitive = True

            # Mark row as non-data if a high-priority heuristic or combination triggers
            should_skip = is_metadata_row or is_repetitive or is_sparse

            if should_skip:
                metadata_rows.append(i)
                sample_text = " | ".join(populated[:4])
                if matched_keyword:
                    keyword = matched_keyword
                elif is_repetitive:
                    keyword = "repetitive values"
                else:
                    keyword = "sparse/metadata"
                additional_skips.append(SkipRowHit(i, keyword, sample_text))

        return metadata_rows, additional_skips

    def _is_numeric(self, s: str) -> bool:
        t = re.sub(r"[,$%\s]", "", s)
        if not t:
            return False
        try:
            float(t)
            return True
        except ValueError:
            return False

    # ── Header row span via vertically-merged cells (openpyxl merged_cells; non-read-only only) ─
    def _detect_header_span(self, ws, header_row_idx0: int, header_col_count: int, read_only: bool) -> int:
        if read_only or header_col_count == 0:
            return 1
        try:
            excel_header_row = header_row_idx0 + 1  # merged-cell ranges are 1-based Excel rows
            span = 1
            for rng in ws.merged_cells.ranges:
                if rng.min_row <= excel_header_row <= rng.max_row and rng.max_row > rng.min_row:
                    if rng.min_col <= header_col_count:  # roughly within the header's column band
                        span = max(span, rng.max_row - rng.min_row + 1)
            return min(span, 3)   # a >3-row header is almost certainly an unrelated merge elsewhere
        except Exception:
            return 1   # style-API quirks must never abort the run — best-effort only

    # ── Color-flag / legend detection — best-effort, bounded to the detected grid's data rows ──
    def _detect_color_flags(self, ws, header_row_idx0: int, total_rows: int) -> list[ColorFlag]:
        try:
            signatures: dict[str, dict] = {}
            excel_start = header_row_idx0 + 2   # first data row, 1-based
            excel_end = min(total_rows, header_row_idx0 + 1 + 2000)  # bound the style scan too
            for r in range(excel_start, excel_end + 1):
                for c in range(1, 40):  # generous column cap; BB grids are rarely wider than this
                    cell = ws.cell(row=r, column=c)
                    if cell.value in (None, ""):
                        continue
                    sig = self._cell_color_signature(cell)
                    if sig is None:
                        continue
                    entry = signatures.setdefault(sig, {"count": 0, "sample": f"{get_column_letter(c)}{r}"})
                    entry["count"] += 1
            flags = [ColorFlag(sig, v["sample"], v["count"]) for sig, v in signatures.items()]
            flags.sort(key=lambda f: -f.occurrences)
            return flags[:8]   # cap — this is a hint list for the Legend sheet, not an audit trail
        except Exception:
            return []   # styling APIs vary by file; never let this abort the analysis

    def _cell_color_signature(self, cell) -> Optional[str]:
        try:
            fill = cell.fill
            if fill is not None and fill.patternType == "solid" and fill.fgColor is not None:
                rgb = getattr(fill.fgColor, "rgb", None)
                if isinstance(rgb, str) and rgb not in ("00000000", "FFFFFFFF"):
                    return f"Fill #{rgb[-6:]}"
        except Exception:
            pass
        try:
            font = cell.font
            if font is not None and font.color is not None:
                rgb = getattr(font.color, "rgb", None)
                if isinstance(rgb, str) and rgb not in ("FF000000", "00000000"):
                    return f"Font #{rgb[-6:]}"
        except Exception:
            pass
        return None

    # ── Agent / title / slug / detect-keys (TemplateProfiler.inferAgent/inferTitle/slugFromFilename) ──
    def _infer_agent(self, rows: Optional[list[list]], header_row: int, override: Optional[str]) -> ProfiledField:
        if override and override.strip():
            return ProfiledField.of(override.strip(), "high", "operator-provided (--agent-bank)")
        if rows:
            limit = header_row if header_row >= 0 else len(rows)
            for i in range(min(limit, len(rows))):
                cells = rows[i]
                for c, cell in enumerate(cells):
                    if _normalize(cell) in AGENT_LABELS:
                        for v in cells[c + 1:]:
                            if v not in (None, ""):
                                return ProfiledField.of(str(v).strip(), "medium",
                                                         f'summary-block label "{str(cells[c]).strip()}"')
        return ProfiledField.of(None, "low", "no agent label found — pass --agent-bank")

    def _infer_title(self, rows: Optional[list[list]]) -> ProfiledField:
        if rows:
            for i in range(min(2, len(rows))):
                for cell in rows[i]:
                    if cell not in (None, "") and len(str(cell).strip()) >= 4:
                        return ProfiledField.of(str(cell).strip(), "medium", f"first non-empty cell in row {i + 1}")
        return ProfiledField.of(None, "low", "no title cell found")

    def _slug_from_filename(self, file_name: str) -> str:
        base = re.sub(r"(?i)\.(xlsx|xls|csv)$", "", file_name or "")
        slug = re.sub(r"[^a-z0-9]+", "-", base.lower())
        slug = re.sub(r"^(agent-bb|bb-template-import|bb-template)-", "", slug)
        slug = slug.strip("-")
        if len(slug) > 50:   # bb_templates.template_slug is VARCHAR(50); trim at a word boundary
            slug = re.sub(r"-+[^-]*$", "", slug[:50])
        return slug

    def _detect_keys(self, slug: str, title: Optional[str]) -> list[str]:
        keys: list[str] = []
        if title and title.strip():
            keys.append(_normalize(title))
        if slug and slug.strip():
            keys.append(slug.replace("-", " "))
        seen = set()
        return [k for k in keys if k and not (k in seen or seen.add(k))]


# ==============================================================================================
#  TemplateBuilder — WorkbookAnalysis -> {"template": BbTemplateRequest-shaped, "analysis": {...}}
# ==============================================================================================
class TemplateBuilder:
    def __init__(self, analysis: WorkbookAnalysis):
        self.a = analysis

    def build_request(self) -> dict:
        """camelCase, exactly BbTemplateRequest's field set — paste as-is into a POST/PUT body."""
        a = self.a
        first = a.tabs[0] if a.tabs else None
        return {
            "templateSlug": a.slug.value,
            "templateName": a.slug.value,   # import convention: no separate display name (see BbTemplateImportService)
            "agentName": a.agent_name.value,
            "templateClass": a.template_class.value,
            "sheetName": first.sheet_name if first else None,
            "headerRowIndex": first.header.value if first else None,
            "autoLearned": True,   # this proposal's origin IS the heuristic tool, not a human author
            "trancheCount": max(1, len(a.tabs)),
            "hasGroupingRows": bool(first and first.groups),
            "hasColorFlags": bool(first and first.color_flags),
            "autoDiscoverTabs": a.auto_discover_tabs.value,
            "summaryRowsAboveHeader": first.header.value if first else 0,
            "summaryRowRange": f"1-{first.header.value}" if first and first.header.value else None,
            "titleRow": 1 if a.title.value else None,
            "titleText": a.title.value,
            "detectKeys": a.detect_keys,
            "legend": [{"style": f.style, "meaning": ""} for f in (first.color_flags if first else [])],
            "notes": [n for tab in a.tabs for n in tab.notes] + self._cross_role_group_caution(),
            "tabs": [self._tab_request(t) for t in a.tabs],
        }

    def _tab_request(self, t: TabAnalysis) -> dict:
        # Only include columns that matched the Field Mapping Dictionary (exclude unmatched headers)
        matched_columns = [f["header"] for f in t.matched_fields]
        return {
            "tabRole": "LP_GRID",   # the only role this tool infers structurally; others are operator-assigned
            "tabSort": t.tab_sort,
            "sheetName": t.sheet_name,
            "sleeveName": t.sheet_name,
            "headerRowIndex": t.header.value,
            "headerRowSpan": t.header_row_span,
            "skipRowKeywords": t.skip_row_keywords,
            "columns": matched_columns,
            "groups": [{"groupSort": i + 1, "headerText": g.header_text, "classification": g.classification}
                       for i, g in enumerate(t.groups)],
        }

    def _cross_role_group_caution(self) -> list[str]:
        """Multiple tabs with groups may require manual review for proper configuration."""
        tabs_with_groups = [t for t in self.a.tabs if t.groups]
        if len(self.a.tabs) > 1 and tabs_with_groups:
            return [
                f"{len(self.a.tabs)} tabs detected with {len(tabs_with_groups)} containing group banners. "
                f"Review tab role assignments to ensure correct group classification."
            ]
        return []

    def build_review_envelope(self) -> dict:
        a = self.a
        review_required: list[str] = []

        def flag(label: str, pf: ProfiledField):
            if pf.confidence in ("medium", "low"):
                review_required.append(f"[{pf.confidence}] {label}: {pf.value!r} — {pf.evidence}")

        flag("slug", a.slug); flag("agent name", a.agent_name); flag("title", a.title)
        flag("template_class", a.template_class); flag("workbook shape", a.workbook_shape)
        for t in a.tabs:
            for g in t.groups:
                if g.confidence != "high":
                    review_required.append(f"[{g.confidence}] tab '{t.sheet_name}' group \"{g.header_text}\" "
                                            f"-> classification \"{g.classification}\"")
            review_required.extend(f"[note] tab '{t.sheet_name}': {n}" for n in t.notes)
        review_required.extend(f"[note] {n}" for n in self._cross_role_group_caution())
        if a.truncated_sheets:
            review_required.append(f"[low] sheet(s) truncated at the row cap during analysis: {a.truncated_sheets} "
                                    f"— re-run with a higher --max-rows-per-sheet if data was cut off")

        return {
            "_readme": (
                "Two sections: 'template' contains the generated template metadata. 'analysis' is this "
                "tool's diagnostic report — confidence/evidence per field, which headers matched "
                "the Field Mapping Dictionary vs. did not, duplicate/continuation tab detection, and a "
                "flat review_required list of everything below high confidence."
            ),
            "template": self.build_request(),
            "analysis": {
                "source_file": a.file_name,
                "analyzed_at": a.analyzed_at,
                "overall_confidence": a.overall_confidence,
                "dictionary_source": a.dictionary_source,
                "dictionary_field_count": a.dictionary_field_count,
                "slug": asdict(a.slug),
                "agent_name": asdict(a.agent_name),
                "title": asdict(a.title),
                "template_class": asdict(a.template_class),
                "workbook_shape": asdict(a.workbook_shape),
                "auto_discover_tabs": asdict(a.auto_discover_tabs),
                "detect_keys": a.detect_keys,
                "non_grid_sheets": a.non_grid_sheets,
                "truncated_sheets": a.truncated_sheets,
                "tabs": [self._tab_analysis_dict(t) for t in a.tabs],
                "review_required": review_required,
            },
        }

    def _tab_analysis_dict(self, t: TabAnalysis) -> dict:
        return {
            "sheet_name": t.sheet_name, "tab_sort": t.tab_sort, "header": asdict(t.header),
            "header_row_span": t.header_row_span, "columns": t.columns,
            "matched_canonical": t.matched_canonical, "matched_fields": t.matched_fields,
            "unmatched_headers": t.unmatched_headers,
            "groups": [asdict(g) for g in t.groups],
            "skip_row_keywords": t.skip_row_keywords,
            "skip_rows_detected": [asdict(h) for h in t.skip_rows_detected],
            "color_flags": [asdict(c) for c in t.color_flags],
            "structure_fingerprint": list(t.structure_fingerprint),
            "duplicate_of": t.duplicate_of,
            "notes": t.notes,
        }

    def write_import_workbook(self, path: Path) -> None:
        """Write the generated template metadata workbook to the specified path."""
        req = self.build_request()
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        ws = wb.create_sheet("Template")
        ws.append(["template_slug", "agent_bank", "template_class", "sheet_name", "header_row_index",
                   "header_row_span", "auto_learned", "tranche_count", "has_grouping_rows",
                   "has_color_flags", "summary_rows_above_header", "summary_row_range",
                   "detect_keys"])
        first_tab = req["tabs"][0] if req["tabs"] else {}
        ws.append([
            req["templateSlug"], req["agentName"], req["templateClass"], req["sheetName"],
            (req["headerRowIndex"] + 1) if req["headerRowIndex"] is not None else None,   # back to 1-based
            first_tab.get("headerRowSpan", 1), req["autoLearned"], req["trancheCount"],
            req["hasGroupingRows"], req["hasColorFlags"], req["summaryRowsAboveHeader"],
            req["summaryRowRange"], ", ".join(req["detectKeys"]),
        ])

        ws = wb.create_sheet("Tabs")
        ws.append(["tab_role", "tab_sort", "sheet_name", "header_row_index",
                   "header_row_span", "skip_row_keywords", "expected_source_headers_json"])
        for t in req["tabs"]:
            ws.append([t["tabRole"], t["tabSort"], t["sheetName"],
                       (t["headerRowIndex"] + 1) if t["headerRowIndex"] is not None else None,
                       t["headerRowSpan"], ", ".join(t["skipRowKeywords"]), json.dumps(t["columns"])])

        ws = wb.create_sheet("Groups")
        ws.append(["tab_role", "group_sort", "header_text", "classification"])
        for t in req["tabs"]:
            for g in t["groups"]:
                ws.append([t["tabRole"], g["groupSort"], g["headerText"], g["classification"]])

        ws = wb.create_sheet("Columns")
        ws.append(["tab_role", "column_sort", "source_header"])
        for t in req["tabs"]:
            for i, col in enumerate(t["columns"], start=1):
                ws.append([t["tabRole"], i, col])

        ws = wb.create_sheet("Legend")
        ws.append(["legend_sort", "style", "meaning"])
        for i, entry in enumerate(req["legend"], start=1):
            ws.append([i, entry["style"], entry["meaning"]])

        ws = wb.create_sheet("Notes")
        ws.append(["note_sort", "note"])
        for i, note in enumerate(req["notes"], start=1):
            ws.append([i, note])

        path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(path)


# ==============================================================================================
#  Console / file summary report
# ==============================================================================================
def _kv(label: str, value: str, width: int = 20) -> str:
    return f"{label:<{width}}: {value}"


def render_summary(analysis: WorkbookAnalysis, envelope: dict) -> str:
    lines = [
        "Excel workbook analysis — summary",
        _kv("generated", analysis.analyzed_at),
        _kv("source file", analysis.file_name),
        _kv("dictionary source", f"{analysis.dictionary_source} ({analysis.dictionary_field_count} canonical fields)"),
        _kv("overall confidence", analysis.overall_confidence),
        _kv("proposed slug", f"{analysis.slug.value}  [{analysis.slug.confidence}]"),
        _kv("proposed agent name", f"{analysis.agent_name.value}  [{analysis.agent_name.confidence}]"),
        _kv("workbook shape", f"{analysis.workbook_shape.value}  [{analysis.workbook_shape.confidence}]"),
        "",
        _kv("sheets scanned", str(len(analysis.tabs) + len(analysis.non_grid_sheets))),
        _kv("  recognized as LP grids", f"{len(analysis.tabs)}  ({[t.sheet_name for t in analysis.tabs]})"),
        _kv("  not recognized", f"{len(analysis.non_grid_sheets)}  ({analysis.non_grid_sheets})"),
    ]
    for t in analysis.tabs:
        lines.append(
            f"\n  tab '{t.sheet_name}' (sort {t.tab_sort}): header row {t.header.value} "
            f"(0-based), {t.matched_canonical} canonical match(es), {len(t.unmatched_headers)} unmatched, "
            f"{len(t.groups)} group(s), {len(t.skip_rows_detected)} skip-row hit(s), "
            f"{len(t.color_flags)} color signature(s)"
        )
        if t.duplicate_of:
            lines.append(f"    ! duplicate structure of '{t.duplicate_of}' — confirm continuation vs. duplicate")
    review = envelope["analysis"]["review_required"]
    lines.append(f"\nreview_required ({len(review)} item(s)):")
    if review:
        lines.extend(f"  - {r}" for r in review)
    else:
        lines.append("  (none — every field cleared high confidence)")
    return "\n".join(lines)


# ==============================================================================================
#  CLI
# ==============================================================================================
def build_arg_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        prog="parse_excel_templates.py",
        description="Analyze an Excel workbook and generate a template metadata file.",
    )
    p.add_argument("input", help="Path to the Excel workbook to analyze (.xlsx or .xls)")
    return p


def _resolve_input(raw: str) -> Path:
    """Resolve input file path (absolute or relative to current working directory)."""
    candidates = [Path(raw)]
    if not Path(raw).is_absolute():
        candidates += [Path.cwd() / raw]
    for c in candidates:
        if c.is_file():
            return c
    raise SystemExit(f"Input workbook not found: {raw}")


def main(argv: Optional[list[str]] = None) -> int:
    for stream in (sys.stdout, sys.stderr):
        try:
            stream.reconfigure(encoding="utf-8", errors="replace")
        except (AttributeError, ValueError):
            pass

    args = build_arg_parser().parse_args(argv)
    input_path = _resolve_input(args.input)

    print(f"Processing: {input_path}")
    dictionary = load_dictionary()

    analyzer = ExcelAnalyzer(dictionary)
    try:
        analysis = analyzer.analyze_workbook(input_path)
    except Exception as e:
        raise SystemExit(f"Failed to analyze '{input_path}': {e}") from e

    builder = TemplateBuilder(analysis)
    envelope = builder.build_review_envelope()
    summary = render_summary(analysis, envelope)
    print("\n" + summary)

    if not analysis.tabs:
        print("\nNo data grid sheet recognized in this workbook.", file=sys.stderr)
        return 1

    slug = analysis.slug.value or "untitled"
    xlsx_path = input_path.parent / f"BB-Template-Import-{slug}.xlsx"
    builder.write_import_workbook(xlsx_path)
    print(f"\nTemplate written: {xlsx_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
